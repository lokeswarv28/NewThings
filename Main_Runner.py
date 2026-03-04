"""
Snowflake Object Lineage Discovery Updated Version On March 02 2026 v2 with some more flexibility
 
Purpose:
- Recursively discovers dependencies for Snowflake views
- Traces lineage from views down to base tables
- Handles circular dependencies and tracks recursion depth
- Analyzes cluster keys and partition keys usage in queries
 
Usage Modes:

Mode 1: Lineage + Cluster Key Check Only
    py Main_Runner4_1_Test3.py --txt source/views.txt --mode lineage-only
                              ↓
    1. Read objects from views.txt
    2. Run Snowflake lineage analysis
    3. Extract cluster/partition keys from lineage
    4. Check which keys are present in query
    5. Save analysis to cluster_key_<queryname>.txt
    6. Display recommendations

Mode 2: Lineage + Cluster Key Check + Query Execution
    py Main_Runner4_1_Test3.py --txt source/views.txt --mode lineage-with-execution --platform snowflake --timeout 5
                              ↓
    1. Read objects and SQL query from views.txt
    2. Run Snowflake lineage analysis
    3. Execute query on selected platform (Snowflake/Databricks)
    4. Monitor execution (kill if exceeds timeout)
    5. Capture Query ID
    6. Fetch GET_QUERY_OPERATOR_STATS → Save to CSV (Snowflake only)
    7. 🔍 AUTOMATICALLY CALL: analyze_expensive_nodes.py
    8. Identify expensive operators, bottlenecks, recommendations
    9. Extract cluster/partition keys from lineage
    10. Check which keys are present in query
    11. Save analysis to cluster_key_<queryname>.txt
    12. Display recommendations

Examples:
    # Lineage and cluster key analysis only
    py Main_Runner4_1_Test3.py --txt source/views.txt --mode lineage-only
    
    # Lineage, cluster key analysis, and query execution on Snowflake
    py Main_Runner4_1_Test3.py --txt source/views.txt --mode lineage-with-execution --platform snowflake --timeout 5
    
    # Lineage, cluster key analysis, and query execution on Databricks
    py Main_Runner4_1_Test3.py --txt source/views.txt --mode lineage-with-execution --platform databricks --timeout 10
"""

# ═══════════════════════════════════════════════════════════════════════════
# Query Performance Analyzer Framework
# Build: LOKV-2026 | Senior Software Engineer
# ═══════════════════════════════════════════════════════════════════════════
 
import snowflake.connector
import os
from dotenv import load_dotenv
import json
from typing import List, Dict, Set, Tuple, Union
import pandas as pd
from teradata_cluster_keys import get_partition_columns
from databricks_conn2 import get_databricks_clustering_info, conn as databricks_conn
import logging
from datetime import datetime
import argparse
import re
import sys
import threading
import time
from databricks import sql as databricks_sql

# Load environment variables from .env file
load_dotenv()

# Note: Logging will be configured in main() after TARGET_OBJECT is defined
logger = None

# Define schemas to exclude from clustering/partition key analysis
EXCLUDED_SCHEMAS = {'UHCCDSECURITY', 'UHCCDSECURITYVIEW', 'UAHSECURITY'}


def get_snowflake_connection():
    """
    Establishes connection to Snowflake using environment variables.
    
    Returns:
        snowflake.connector.connection: Active Snowflake connection
    """ 
    return snowflake.connector.connect(
        account=os.getenv("s_account"),
        user=os.getenv("s_user"),
        authenticator="externalbrowser",
        warehouse=os.getenv("s_warehouse"),
        database=os.getenv("s_database"),
        schema=os.getenv("s_schema"),
        role=os.getenv("s_role")
    )


def execute_databricks_query_with_timeout(conn, query, timeout_minutes=10):
    """
    Execute a query in Databricks with timeout. Cancels the query if it exceeds the timeout.
    
    Args:
        conn: Databricks connection
        query: SQL query to execute
        timeout_minutes: Maximum execution time in minutes (default 10)
        
    Returns:
        tuple: (query_id, was_killed, error_message, query_metrics)
    """
    query_id = None
    was_killed = False
    error_message = None
    result = {'finished': False, 'error': None, 'query_id': None, 'cursor': None}
    
    def execute_query():
        try:
            cursor = conn.cursor()
            result['cursor'] = cursor
            
            # Log what we're about to execute
            if logger:
                logger.info(f"=== EXECUTING USER QUERY IN THREAD ===")
                logger.info(f"Query preview: {query[:200]}...")
            print(f"\n🔍 DEBUG: About to execute query in thread...")
            print(f"   Query preview: {query[:100]}...")
            
            cursor.execute("USE CATALOG prod_uhccd;")
            
            if logger:
                logger.info(f"Catalog set. Now executing main query...")
            print(f"✅ Catalog set. Now executing main query...")
            
            cursor.execute(query)
            
            if logger:
                logger.info(f"Query execution completed in thread")
            print(f"✅ Query execution completed in thread")
            
            result['finished'] = True
            
            cursor.close()
            
        except Exception as e:
            result['error'] = str(e)
            result['finished'] = True
            if result['cursor']:
                try:
                    result['cursor'].close()
                except:
                    pass
    
    # Start query execution in a separate thread
    query_thread = threading.Thread(target=execute_query)
    query_thread.daemon = True
    
    # Record start time BEFORE starting thread to include all time
    execution_start_time = time.time()
    timeout_seconds = timeout_minutes * 60
    
    query_thread.start()
    
    # Wait longer for query to start and be recorded in query_history
    print(f"\n⏳ Waiting 8 seconds for query to be recorded in query_history...")
    if logger:
        logger.info("Waiting 8 seconds for query to be recorded in query_history")
    time.sleep(8)
    
    # Try to capture Query ID from Databricks using system_catalog.query_history
    # CRITICAL: Match by query text similarity to avoid capturing metadata queries
    # 
    # FIX: Previously used keyword filters (NOT LIKE '%DESCRIBE DETAIL%') which was brittle
    #      and missed new metadata query types, causing wrong Query IDs to be captured.
    # 
    # NEW APPROACH: Extract 50-char signature from user query and match against query_history
    #               This ensures we ONLY capture the exact user query, not metadata queries.
    # 
    max_retries = 3
    retry_count = 0
    
    while retry_count < max_retries and not query_id:
        try:
            temp_cursor = conn.cursor()
            
            # Extract a unique signature from the user's query (first 150 chars, normalized)
            # Increased from 50 to 150 chars for better uniqueness when queries have similar starts
            query_signature = ' '.join(query.strip().split())[:150].upper()
            
            if logger:
                logger.info(f"Attempt {retry_count + 1}/{max_retries}: Looking for query with signature: {query_signature}...")
            print(f"\n🔍 Attempt {retry_count + 1}/{max_retries}: Searching for Query ID...")
            print(f"   Signature (first 80 chars): {query_signature[:80]}...")
            
            # Query for the most recent query that matches the signature
            # This is more reliable than filtering out metadata queries
            temp_cursor.execute("""
                SELECT statement_id, statement_type, execution_status, start_time, end_time, statement_text
                FROM system_catalog.query.query_history
                WHERE executed_by = current_user() 
                  AND start_time >= date_sub(current_timestamp(), 60)
                  AND statement_type = 'SELECT'
                ORDER BY start_time DESC 
                LIMIT 15
            """)
            query_results = temp_cursor.fetchall()
            
            if logger:
                logger.info(f"Retrieved {len(query_results)} candidate queries from query_history")
            print(f"   Found {len(query_results)} candidate SELECT queries in history")
            
            # Find the query that best matches our signature
            best_match = None
            for idx, row in enumerate(query_results, 1):
                statement_id, statement_type, execution_status, start_time, end_time, statement_text = row
                if statement_text:
                    # Normalize and extract signature from candidate (using 150 chars now)
                    candidate_signature = ' '.join(statement_text.strip().split())[:150].upper()
                    
                    if logger:
                        logger.debug(f"  Candidate {idx}: {statement_id[:20]}... | Sig match: {candidate_signature == query_signature} | Signature: {candidate_signature[:60]}...")
                    
                    print(f"   Candidate {idx}: {statement_id} | Match: {candidate_signature == query_signature}")
                    
                    # Check for match
                    if candidate_signature == query_signature:
                        best_match = row
                        if logger:
                            logger.info(f"[OK] Found exact match: {statement_id}")
                        print(f"   ✅ Match found: {statement_id}")
                        break
            
            if best_match:
                query_id = best_match[0]
                query_text = best_match[5] if len(best_match) > 5 else None
                if logger:
                    logger.info(f"[OK] Databricks Query ID (statement_id) captured: {query_id}")
                    if query_text:
                        logger.info(f"  Query text preview: {query_text[:150]}...")
                print(f"\n🆔 Databricks Query ID (statement_id) captured: {query_id}")
                if query_text:
                    print(f"   Query preview: {query_text[:100]}...")
            else:
                if logger:
                    logger.warning(f"No match found in attempt {retry_count + 1}")
                print(f"   ⚠️  No match found")
                
            temp_cursor.close() 
            
        except Exception as e:
            if logger:
                logger.warning(f"Error in attempt {retry_count + 1}: {str(e)}")
            print(f"\n⚠️  Error in attempt {retry_count + 1}: {str(e)}")
        
        # If no match found and retries remaining, wait and try again
        if not query_id and retry_count < max_retries - 1:
            wait_time = 5
            if logger:
                logger.info(f"Waiting {wait_time} seconds before retry...")
            print(f"   Waiting {wait_time} seconds before retry...")
            time.sleep(wait_time)
        
        retry_count += 1
    
    # Display captured query ID
    if query_id:
        if logger:
            logger.info(f"=======================================================")
            logger.info(f"  DATABRICKS QUERY ID: {query_id}")
            logger.info(f"=======================================================")
        print(f"\n{'='*80}")
        print(f"🆔 ACTIVE DATABRICKS QUERY ID: {query_id}")
        print(f"{'='*80}\n")
    
    # Poll for query status (check if QUEUED → RUNNING)
    current_status = None
    queued_start_time = time.time()
    max_queued_wait = 300  # Maximum 5 minutes to wait for query to leave QUEUED state
    
    if query_id:
        print(f"📊 Checking query status...")
        while True:
            try:
                status_cursor = conn.cursor()
                status_cursor.execute(f"""
                    SELECT execution_status 
                    FROM system_catalog.query.query_history 
                    WHERE statement_id = '{query_id}'
                """)
                status_row = status_cursor.fetchone()
                status_cursor.close()
                
                if status_row:
                    current_status = status_row[0]
                    
                    if current_status == 'QUEUED':
                        queued_elapsed = time.time() - queued_start_time
                        if logger:
                            logger.warning(f"Query is QUEUED (waiting for resources) - {int(queued_elapsed)}s elapsed")
                        print(f"⏳ Query is QUEUED (waiting for resources) - {int(queued_elapsed)}s elapsed...")
                        
                        # Check if we've been queued too long
                        if queued_elapsed > max_queued_wait:
                            if logger:
                                logger.error(f"Query has been QUEUED for {int(queued_elapsed)}s. Warehouse may be at capacity.")
                            print(f"\n⚠️  WARNING: Query has been QUEUED for {int(queued_elapsed)}s")
                            print(f"   Warehouse may be at capacity or starting up.")
                            print(f"   Continuing to monitor...\n")
                        
                        time.sleep(5)  # Check every 5 seconds while queued
                        continue
                        
                    elif current_status == 'RUNNING':
                        if logger:
                            logger.info(f"[OK] Query status changed to RUNNING")
                        print(f"✅ Query is now RUNNING (monitoring for timeout)")
                        break
                        
                    elif current_status in ['FINISHED', 'FAILED', 'CANCELED']:
                        if logger:
                            logger.info(f"Query already completed with status: {current_status}")
                        print(f"✅ Query completed with status: {current_status}")
                        break
                        
                    else:
                        # Unknown status, break and proceed with monitoring
                        if logger:
                            logger.warning(f"Unknown query status: {current_status}")
                        print(f"⚠️  Unknown query status: {current_status}")
                        break
                else:
                    # No status found yet, wait and retry
                    time.sleep(2)
                    continue
                    
            except Exception as e:
                if logger:
                    logger.warning(f"Could not check query status: {str(e)}")
                print(f"⚠️  Could not check query status, proceeding with monitoring...")
                break
    
    # Monitor execution time with forced timeout
    last_status_check = time.time()
    status_check_interval = 10  # Check status every 10 seconds
    
    print(f"⏱️  Monitoring query execution (timeout: {timeout_minutes} minutes)...")
    
    while query_thread.is_alive():
        elapsed = time.time() - execution_start_time
        
        # CRITICAL: Check timeout FIRST before any other checks
        if elapsed > timeout_seconds:
            # Timeout reached - cancel the query
            was_killed = True
            
            if logger:
                logger.warning(f"Query execution exceeded {timeout_minutes} minutes. Cancelling query...")
                logger.warning(f"Query ID to cancel: {query_id}")
            print(f"\n⚠️  Query execution exceeded {timeout_minutes} minutes. Cancelling query...")
            print(f"🆔 Query ID to cancel: {query_id}")
            
            # Cancel the query in Databricks
            if query_id:
                try:
                    cancel_cursor = conn.cursor()
                    cancel_cursor.execute(f"CANCEL QUERY '{query_id}'")
                    if logger:
                        logger.info(f"[OK] Query cancelled successfully. Query ID: {query_id}")
                    print(f"✅ Query cancelled successfully. Query ID: {query_id}")
                    cancel_cursor.close()
                except Exception as e:
                    error_message = f"Failed to cancel query: {str(e)}"
                    if logger:
                        logger.error(error_message)
                    print(f"❌ {error_message}")
            else:
                error_message = "Could not retrieve Query ID to cancel the query"
                if logger:
                    logger.error(error_message)
                print(f"❌ {error_message}")
            
            # Force close cursor if still open
            if result['cursor']:
                try:
                    result['cursor'].close()
                    print(f"🔌 Forced cursor close")
                except Exception as e:
                    print(f"⚠️  Could not force close cursor: {e}")
            
            # Don't wait for thread - break immediately
            print(f"⏹️  Breaking out of monitoring loop (thread will be terminated)")
            break
        
        # Periodically check and display status
        if query_id and (time.time() - last_status_check) >= status_check_interval:
            try:
                status_cursor = conn.cursor()
                status_cursor.execute(f"""
                    SELECT execution_status 
                    FROM system_catalog.query.query_history 
                    WHERE statement_id = '{query_id}'
                """)
                status_row = status_cursor.fetchone()
                status_cursor.close()
                
                if status_row:
                    new_status = status_row[0]
                    if new_status != current_status:
                        if logger:
                            logger.info(f"Query status changed: {current_status} → {new_status}")
                        print(f"🔄 Query status: {current_status} → {new_status} (elapsed: {int(elapsed)}s)")
                        current_status = new_status
            except Exception as e:
                pass  # Silently continue if status check fails
            
            last_status_check = time.time()
        
        time.sleep(1)  # Check every second
    
    # Wait for thread to finish (but only briefly since we may have cancelled it)
    if was_killed:
        print(f"⏳ Waiting up to 10 seconds for thread to terminate after cancellation...")
        query_thread.join(timeout=10)
        if query_thread.is_alive():
            print(f"⚠️  WARNING: Thread still alive after cancellation. Query may still be running on Databricks.")
            print(f"   You may need to manually cancel Query ID: {query_id}")
    else:
        print(f"⏳ Waiting for query thread to complete...")
        query_thread.join(timeout=5)
    
    # Check for errors
    if not was_killed:
        if result['error']:
            error_message = result['error']
            if logger:
                logger.error(f"Query error: {error_message}")
            print(f"❌ Query error: {error_message}")
        else:
            if logger:
                logger.info(f"Query completed. Query ID: {query_id}")
            print(f"\n✅ Query completed. Query ID: {query_id}")
    
    # Now fetch query metrics from system_catalog.query.query_history
    query_metrics = None
    if query_id:
        try:
            if logger:
                logger.info(f"Fetching query metrics for query_id: {query_id}")
            print(f"\n📊 Fetching query metrics for query_id: {query_id}...")
            
            metrics_cursor = conn.cursor()
            metrics_query = f"""
                SELECT 
                    Execution_status, 
                    Statement_type, 
                    execution_duration_ms, 
                    start_time, 
                    end_time, 
                    read_bytes, 
                    read_files, 
                    read_partitions, 
                    read_rows, 
                    produced_rows, 
                    written_bytes, 
                    written_files, 
                    written_rows, 
                    spilled_local_bytes, 
                    shuffle_read_bytes, 
                    workspace_name
                FROM system_catalog.query.query_history 
                WHERE statement_id = '{query_id}'
            """
            metrics_cursor.execute(metrics_query)
            metrics_row = metrics_cursor.fetchone()
            
            if metrics_row:
                # Create a dictionary with column names and values
                query_metrics = {
                    'execution_status': metrics_row[0],
                    'statement_type': metrics_row[1],
                    'execution_duration_ms': metrics_row[2],
                    'start_time': metrics_row[3],
                    'end_time': metrics_row[4],
                    'read_bytes': metrics_row[5],
                    'read_files': metrics_row[6],
                    'read_partitions': metrics_row[7],
                    'read_rows': metrics_row[8],
                    'produced_rows': metrics_row[9],
                    'written_bytes': metrics_row[10],
                    'written_files': metrics_row[11],
                    'written_rows': metrics_row[12],
                    'spilled_local_bytes': metrics_row[13],
                    'shuffle_read_bytes': metrics_row[14],
                    'workspace_name': metrics_row[15]
                }
                
                if logger:
                    logger.info(f"[OK] Query metrics retrieved successfully")
                    logger.info(f"  Status: {query_metrics['execution_status']}")
                    logger.info(f"  Type: {query_metrics['statement_type']}")
                    logger.info(f"  Duration: {query_metrics['execution_duration_ms']} ms")
                    logger.info(f"  Rows Read: {query_metrics['read_rows']}")
                    logger.info(f"  Rows Produced: {query_metrics['produced_rows']}")
                
                print(f"✅ Query metrics retrieved successfully")
                
                # Color-code status display
                status_emoji = "✅" if query_metrics['execution_status'] == 'FINISHED' else \
                               "⏳" if query_metrics['execution_status'] == 'QUEUED' else \
                               "🔄" if query_metrics['execution_status'] == 'RUNNING' else \
                               "❌" if query_metrics['execution_status'] == 'FAILED' else "⚠️"
                
                print(f"\n{'='*80}")
                print(f"📊 FINAL QUERY METRICS")
                print(f"{'='*80}")
                print(f"{status_emoji} Status: {query_metrics['execution_status']}")
                print(f"   Type: {query_metrics['statement_type']}")
                print(f"   Duration: {query_metrics['execution_duration_ms']} ms")
                print(f"   Start Time: {query_metrics['start_time']}")
                print(f"   End Time: {query_metrics['end_time']}")
                print(f"   Rows Read: {query_metrics['read_rows']}")
                print(f"   Rows Produced: {query_metrics['produced_rows']}")
                print(f"   Files Read: {query_metrics['read_files']}")
                print(f"   Bytes Read: {query_metrics['read_bytes']}")
                print(f"{'='*80}")
            else:
                if logger:
                    logger.warning(f"No metrics found for query_id: {query_id}")
                print(f"⚠️  No metrics found for query_id: {query_id}")
            
            metrics_cursor.close()
            
        except Exception as e:
            if logger:
                logger.error(f"Failed to retrieve query metrics: {str(e)}")
            print(f"❌ Failed to retrieve query metrics: {str(e)}")
    else:
        if logger:
            logger.warning("Cannot fetch metrics - query_id is None")
        print(f"⚠️  Cannot fetch metrics - query_id is None")
    
    return query_id, was_killed, error_message, query_metrics


def execute_query_with_timeout(cursor, query, timeout_minutes=25):
    """
    Execute a query with timeout. Kills the query if it exceeds the timeout.
    
    Args:
        cursor: Snowflake cursor
        query: SQL query to execute
        timeout_minutes: Maximum execution time in minutes (default 25)
        
    Returns:
        tuple: (query_id, was_killed, error_message, query_metrics)
    """
    query_id = None
    was_killed = False
    error_message = None
    result = {'finished': False, 'error': None, 'query_id': None}
    
    # Get current session ID to track queries
    try:
        cursor.execute("SELECT CURRENT_SESSION()")
        session_id = cursor.fetchone()[0]
        if logger:
            logger.info(f"Current session ID: {session_id}")
    except Exception as e:
        session_id = None
        if logger:
            logger.warning(f"Could not get session ID: {str(e)}")
    
    def execute_query():
        try:
            cursor.execute(query)
            result['finished'] = True
            # Try multiple methods to get query ID
            query_id_captured = None
            
            # Method 1: cursor.sfqid attribute
            if hasattr(cursor, 'sfqid') and cursor.sfqid:
                query_id_captured = cursor.sfqid
                if logger:
                    logger.info(f"[OK] Method 1: Query ID from cursor.sfqid: {query_id_captured}")
            
            # Method 2: LAST_QUERY_ID() on same cursor
            if not query_id_captured:
                try:
                    cursor.execute("SELECT LAST_QUERY_ID()")
                    query_id_captured = cursor.fetchone()[0]
                    if logger:
                        logger.info(f"[OK] Method 2: Query ID from LAST_QUERY_ID(): {query_id_captured}")
                except Exception as e:
                    if logger:
                        logger.warning(f"Method 2 failed: {str(e)}")
            
            result['query_id'] = query_id_captured
            
        except Exception as e:
            result['error'] = str(e)
            result['finished'] = True
            # Try to get query ID even on error
            try:
                if hasattr(cursor, 'sfqid') and cursor.sfqid:
                    result['query_id'] = cursor.sfqid
                elif hasattr(cursor, '_query_result_format') and hasattr(cursor, '_query'):
                    # Try another internal attribute
                    result['query_id'] = getattr(cursor, '_last_query_id', None)
            except:
                pass
    
    # Start query execution in a separate thread
    query_thread = threading.Thread(target=execute_query)
    query_thread.daemon = True
    query_thread.start()
    
    # Wait for query to start and attempt to capture Query ID
    time.sleep(5)
    
    # Try to get Query ID from execution thread first (Methods 1 & 2)
    if result.get('query_id'):
        query_id = result['query_id']
        if logger:
            logger.info(f"[OK] Query ID captured from execution thread (Method 1 or 2): {query_id}")
        print(f"\n🆔 Query ID captured from execution thread: {query_id}")
    
    # Method 3: Query session history if Methods 1 & 2 failed
    if not query_id and session_id:
        try:
            temp_cursor = cursor.connection.cursor()
            # Get most recent query from this session (excluding metadata queries)
            temp_cursor.execute(f"""
                SELECT QUERY_ID 
                FROM TABLE(INFORMATION_SCHEMA.QUERY_HISTORY_BY_SESSION({session_id}))
                WHERE QUERY_TEXT NOT LIKE '%CURRENT_SESSION%'
                  AND QUERY_TEXT NOT LIKE '%QUERY_HISTORY%'
                  AND QUERY_TEXT NOT LIKE '%SYSTEM$CANCEL_QUERY%'
                ORDER BY START_TIME DESC 
                LIMIT 1
            """)
            result_row = temp_cursor.fetchone()
            if result_row and result_row[0]:
                query_id = result_row[0]
                if logger:
                    logger.info(f"[OK] Method 3: Query ID from session history: {query_id}")
                print(f"\n🆔 Query ID captured from session history (Method 3): {query_id}")
            temp_cursor.close()
        except Exception as e:
            if logger:
                logger.warning(f"Method 3 failed: {str(e)}")
    
    # Display captured query ID
    if query_id:
        if logger:
            logger.info(f"=======================================================")
            logger.info(f"  SNOWFLAKE QUERY ID: {query_id}")
            logger.info(f"=======================================================")
        print(f"\n{'='*80}")
        print(f"🆔 ACTIVE SNOWFLAKE QUERY ID: {query_id}")
        print(f"{'='*80}\n")
    
    # Poll for query status (check if QUEUED → RUNNING)
    current_status = None
    queued_start_time = time.time()
    max_queued_wait = 300  # Maximum 5 minutes to wait for query to leave QUEUED state
    
    if query_id:
        print(f"📊 Checking query status...")
        while True:
            try:
                status_cursor = cursor.connection.cursor()
                status_cursor.execute(f"""
                    SELECT EXECUTION_STATUS 
                    FROM TABLE(INFORMATION_SCHEMA.QUERY_HISTORY())
                    WHERE QUERY_ID = '{query_id}'
                    LIMIT 1
                """)
                status_row = status_cursor.fetchone()
                status_cursor.close()
                
                if status_row:
                    current_status = status_row[0]
                    
                    if current_status == 'QUEUED':
                        queued_elapsed = time.time() - queued_start_time
                        if logger:
                            logger.warning(f"Query is QUEUED (waiting for warehouse) - {int(queued_elapsed)}s elapsed")
                        print(f"⏳ Query is QUEUED (waiting for warehouse) - {int(queued_elapsed)}s elapsed...")
                        
                        # Check if we've been queued too long
                        if queued_elapsed > max_queued_wait:
                            if logger:
                                logger.error(f"Query has been QUEUED for {int(queued_elapsed)}s. Warehouse may be suspended or at capacity.")
                            print(f"\n⚠️  WARNING: Query has been QUEUED for {int(queued_elapsed)}s")
                            print(f"   Warehouse may be suspended or at capacity.")
                            print(f"   Continuing to monitor...\n")
                        
                        time.sleep(5)  # Check every 5 seconds while queued
                        continue
                        
                    elif current_status == 'RUNNING':
                        if logger:
                            logger.info(f"[OK] Query status changed to RUNNING")
                        print(f"✅ Query is now RUNNING (monitoring for timeout)")
                        break
                        
                    elif current_status in ['SUCCESS', 'FAILED_WITH_ERROR', 'FAILED_WITH_INCIDENT', 'ABORTED', 'BLOCKED']:
                        if logger:
                            logger.info(f"Query already completed with status: {current_status}")
                        print(f"✅ Query completed with status: {current_status}")
                        break
                        
                    else:
                        # Unknown status, break and proceed with monitoring
                        if logger:
                            logger.warning(f"Unknown query status: {current_status}")
                        print(f"⚠️  Unknown query status: {current_status}")
                        break
                else:
                    # No status found yet, wait and retry
                    time.sleep(2)
                    continue
                    
            except Exception as e:
                if logger:
                    logger.warning(f"Could not check query status: {str(e)}")
                print(f"⚠️  Could not check query status, proceeding with monitoring...")
                break
    
    # Monitor execution time
    start_time = time.time()
    timeout_seconds = timeout_minutes * 60
    last_status_check = time.time()
    status_check_interval = 10  # Check status every 10 seconds
    
    while query_thread.is_alive():
        elapsed = time.time() - start_time
        
        # Check periodically if we got query_id from the thread
        if not query_id and result.get('query_id'):
            query_id = result['query_id']
            if logger:
                logger.info(f"[OK] Query ID updated from thread during monitoring: {query_id}")
        
        # Periodically check and display status
        if query_id and (time.time() - last_status_check) >= status_check_interval:
            try:
                status_cursor = cursor.connection.cursor()
                status_cursor.execute(f"""
                    SELECT EXECUTION_STATUS 
                    FROM TABLE(INFORMATION_SCHEMA.QUERY_HISTORY())
                    WHERE QUERY_ID = '{query_id}'
                    LIMIT 1
                """)
                status_row = status_cursor.fetchone()
                status_cursor.close()
                
                if status_row:
                    new_status = status_row[0]
                    if new_status != current_status:
                        if logger:
                            logger.info(f"Query status changed: {current_status} → {new_status}")
                        print(f"🔄 Query status: {current_status} → {new_status} (elapsed: {int(elapsed)}s)")
                        current_status = new_status
            except Exception as e:
                pass  # Silently continue if status check fails
            
            last_status_check = time.time()
        
        if elapsed > timeout_seconds:
            # Timeout reached - kill the query
            was_killed = True
            
            # Final attempt to get query_id from result if still not captured
            if not query_id and result.get('query_id'):
                query_id = result['query_id']
            
            if logger:
                logger.warning(f"Query execution exceeded {timeout_minutes} minutes. Killing query...")
                logger.warning(f"Query ID to kill: {query_id}")
            print(f"\n⚠️  Query execution exceeded {timeout_minutes} minutes. Killing query...")
            print(f"🆔 Query ID to kill: {query_id}")
            
            # Kill the query using the query_id
            if query_id:
                try:
                    temp_cursor = cursor.connection.cursor()
                    temp_cursor.execute(f"SELECT SYSTEM$CANCEL_QUERY('{query_id}')")
                    if logger:
                        logger.info(f"[OK] Query killed successfully. Query ID: {query_id}")
                    print(f"✅ Query killed successfully. Query ID: {query_id}")
                    temp_cursor.close()
                except Exception as e:
                    error_message = f"Failed to kill query: {str(e)}"
                    if logger:
                        logger.error(error_message)
                    print(f"❌ {error_message}")
            else:
                error_message = "Could not retrieve Query ID to kill the query"
                if logger:
                    logger.error(error_message)
                print(f"❌ {error_message}")
            
            break
        
        time.sleep(1)  # Check every second
    
    # Wait for thread to finish
    query_thread.join(timeout=5)
    
    # Get query ID from result if available
    if not query_id and result.get('query_id'):
        query_id = result['query_id']
        if logger:
            logger.info(f"Query ID retrieved from execution thread: {query_id}")
    
    # Get query ID if not already retrieved
    if not query_id and not was_killed:
        try:
            temp_cursor = cursor.connection.cursor()
            temp_cursor.execute("SELECT LAST_QUERY_ID()")
            query_id = temp_cursor.fetchone()[0]
            temp_cursor.close()
            if logger:
                logger.info(f"Query ID retrieved after completion: {query_id}")
            
            # Log and print QueryID for successful completion
            
            # Log and print QueryID for successful completion
            if logger:
                logger.info(f"Query completed. Query ID: {query_id}")
            print(f"\n✅ Query completed. Query ID: {query_id}")
            
            if result['error']:
                error_message = result['error']
                if logger:
                    logger.error(f"Query error for Query ID {query_id}: {error_message}")
                print(f"❌ Query error for Query ID {query_id}: {error_message}")
        except Exception as e:
            error_msg = f"Failed to get query ID: {str(e)}"
            if logger:
                logger.error(error_msg)
            # Don't overwrite existing error_message if query failed
            if not error_message:
                error_message = error_msg
    
    # Final check - if we still don't have query_id but query finished, log warning
    if not query_id:
        warning_msg = "Warning: Query ID could not be captured. Query may still be running in Snowflake."
        if logger:
            logger.warning(warning_msg)
        print(f"\n⚠️  {warning_msg}")
    
    # Fetch query metrics from QUERY_HISTORY
    query_metrics = None
    if query_id:
        try:
            if logger:
                logger.info(f"Fetching query metrics for query_id: {query_id}")
            print(f"\n📊 Fetching query metrics for query_id: {query_id}...")
            
            metrics_cursor = cursor.connection.cursor()
            metrics_query = f"""
                SELECT 
                    EXECUTION_STATUS,
                    QUERY_TYPE,
                    TOTAL_ELAPSED_TIME as execution_duration_ms,
                    START_TIME,
                    END_TIME,
                    BYTES_SCANNED as read_bytes,
                    BYTES_WRITTEN as written_bytes,
                    ROWS_PRODUCED as produced_rows,
                    WAREHOUSE_NAME,
                    DATABASE_NAME,
                    SCHEMA_NAME
                FROM TABLE(INFORMATION_SCHEMA.QUERY_HISTORY())
                WHERE QUERY_ID = '{query_id}'
                LIMIT 1
            """
            metrics_cursor.execute(metrics_query)
            metrics_row = metrics_cursor.fetchone()
            
            if metrics_row:
                # Create a dictionary with column names and values
                query_metrics = {
                    'execution_status': metrics_row[0],
                    'query_type': metrics_row[1],
                    'execution_duration_ms': metrics_row[2],
                    'start_time': metrics_row[3],
                    'end_time': metrics_row[4],
                    'read_bytes': metrics_row[5],
                    'written_bytes': metrics_row[6],
                    'produced_rows': metrics_row[7],
                    'warehouse_name': metrics_row[8],
                    'database_name': metrics_row[9],
                    'schema_name': metrics_row[10]
                }
                
                if logger:
                    logger.info(f"[OK] Query metrics retrieved successfully")
                    logger.info(f"  Status: {query_metrics['execution_status']}")
                    logger.info(f"  Type: {query_metrics['query_type']}")
                    logger.info(f"  Duration: {query_metrics['execution_duration_ms']} ms")
                    logger.info(f"  Rows Produced: {query_metrics['produced_rows']}")
                
                print(f"✅ Query metrics retrieved successfully")
                
                # Color-code status display
                status_emoji = "✅" if query_metrics['execution_status'] == 'SUCCESS' else \
                               "⏳" if query_metrics['execution_status'] == 'QUEUED' else \
                               "🔄" if query_metrics['execution_status'] == 'RUNNING' else \
                               "❌" if 'FAILED' in query_metrics['execution_status'] else "⚠️"
                
                print(f"\n{'='*80}")
                print(f"📊 FINAL QUERY METRICS (SNOWFLAKE)")
                print(f"{'='*80}")
                print(f"{status_emoji} Status: {query_metrics['execution_status']}")
                print(f"   Type: {query_metrics['query_type']}")
                print(f"   Duration: {query_metrics['execution_duration_ms']} ms")
                print(f"   Start Time: {query_metrics['start_time']}")
                print(f"   End Time: {query_metrics['end_time']}")
                print(f"   Rows Produced: {query_metrics['produced_rows']}")
                print(f"   Bytes Read: {query_metrics['read_bytes']}")
                print(f"   Bytes Written: {query_metrics['written_bytes']}")
                print(f"   Warehouse: {query_metrics['warehouse_name']}")
                print(f"   Database: {query_metrics['database_name']}")
                print(f"   Schema: {query_metrics['schema_name']}")
                print(f"{'='*80}")
            else:
                if logger:
                    logger.warning(f"No metrics found for query_id: {query_id}")
                print(f"⚠️  No metrics found for query_id: {query_id}")
            
            metrics_cursor.close()
            
        except Exception as e:
            if logger:
                logger.error(f"Failed to retrieve query metrics: {str(e)}")
            print(f"❌ Failed to retrieve query metrics: {str(e)}")
    else:
        if logger:
            logger.warning("Cannot fetch metrics - query_id is None")
        print(f"⚠️  Cannot fetch metrics - query_id is None")
    
    return query_id, was_killed, error_message, query_metrics


def get_query_profile(cursor, query_id, output_folder="Output"):
    """
    Get query profile using GET_QUERY_OPERATOR_STATS.
    
    Args:
        cursor: Snowflake cursor
        query_id: Query ID to profile
        output_folder: Folder to save CSV (default: Output)
        
    Returns:
        pd.DataFrame: Query profile data
    """
    profile_query = f"""
    WITH ops AS (
      SELECT *
      FROM TABLE(GET_QUERY_OPERATOR_STATS('{query_id}'))
    )
    SELECT
      OPERATOR_ID,
      OPERATOR_TYPE,
      OPERATOR_NAME,
      STEP_ID,
      EXECUTION_TIME_BREAKDOWN:overall_percentage::FLOAT          AS pct_total_time,
      EXECUTION_TIME_BREAKDOWN:processing::FLOAT                  AS pct_processing,
      EXECUTION_TIME_BREAKDOWN:local_disk_io::FLOAT               AS pct_local_io,
      EXECUTION_TIME_BREAKDOWN:remote_disk_io::FLOAT              AS pct_remote_io,
      EXECUTION_TIME_BREAKDOWN:network_communication::FLOAT       AS pct_network,
      OPERATOR_STATISTICS:input_rows::NUMBER                      AS input_rows,
      OPERATOR_STATISTICS:output_rows::NUMBER                     AS rows_produced,
      OPERATOR_STATISTICS:io:bytes_scanned::NUMBER                AS bytes_scanned,
      OPERATOR_STATISTICS:io:bytes_written::NUMBER                AS bytes_written,
      OPERATOR_STATISTICS:spilling:bytes_spilled_local_storage::NUMBER  AS spilled_local_bytes,
      OPERATOR_STATISTICS:spilling:bytes_spilled_remote_storage::NUMBER AS spilled_remote_bytes
    FROM ops
    ORDER BY pct_total_time DESC NULLS LAST
    LIMIT 10
    """
    
    try:
        if logger:
            logger.info(f"Fetching query profile for Query ID: {query_id}")
        print(f"\n📊 Fetching query profile for Query ID: {query_id}...")
        
        cursor.execute(profile_query)
        results = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        
        df = pd.DataFrame(results, columns=columns)
        
        if logger:
            logger.info(f"Query profile retrieved: {len(df)} operators")
        print(f"✅ Query profile retrieved: {len(df)} operators")
        
        return df
    except Exception as e:
        error_msg = f"Failed to get query profile: {str(e)}"
        if logger:
            logger.error(error_msg)
        print(f"❌ {error_msg}")
        return None


def process_query_file(txt_file_path, timeout_minutes=25, platform='snowflake'):
    """
    Process a .txt file containing SQL query, execute it, and generate profile if timeout occurs.
    
    Args:
        txt_file_path: Path to the .txt file containing SQL query
        timeout_minutes: Maximum execution time in minutes
        platform: 'snowflake' or 'databricks'
    """
    # Extract base filename for output files
    base_filename = os.path.splitext(os.path.basename(txt_file_path))[0]
    
    # Setup logging
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    logs_dir = "logs"
    os.makedirs(logs_dir, exist_ok=True)
    log_filename = os.path.join(logs_dir, f"{base_filename}_query_exec_{timestamp}.log")
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_filename, encoding='utf-8'),
            logging.StreamHandler()
        ],
        force=True
    )
    global logger
    logger = logging.getLogger(__name__)
    
    print("\n" + "="*80)
    print("🚀 QUERY EXECUTION AND PROFILING")
    print("   Build: LOKV-2026 | Senior Software Engineer")
    print("="*80)
    print(f"📄 File: {txt_file_path}")
    print(f"⏱️  Timeout: {timeout_minutes} minutes")
    print(f"☁️  Platform: {platform.upper()}")
    print("="*80 + "\n")
    
    logger.info("="*80)
    logger.info("Query Execution and Profiling Started")
    logger.info(f"File: {txt_file_path}")
    logger.info(f"Timeout: {timeout_minutes} minutes")
    logger.info(f"Platform: {platform.upper()}")
    logger.info("="*80)
    
    # Read the query from file
    try:
        with open(txt_file_path, 'r') as f:
            query = f.read().strip()
        logger.info(f"Query loaded from file ({len(query)} characters)")
        print(f"✅ Query loaded from file ({len(query)} characters)\n")
    except Exception as e:
        error_msg = f"Failed to read file: {str(e)}"
        logger.error(error_msg)
        print(f"❌ {error_msg}")
        return
    
    # Platform-specific connection and execution
    if platform.lower() == 'databricks':
        # Connect to Databricks
        try:
            DATABRICKS_HOST = os.getenv("DATABRICKS_HOST")
            DATABRICKS_HTTP_PATH = os.getenv("DATABRICKS_HTTP_PATH")
            DATABRICKS_TOKEN = os.getenv("DATABRICKS_TOKEN")
            
            if not DATABRICKS_HOST or not DATABRICKS_HTTP_PATH or not DATABRICKS_TOKEN:
                raise Exception("Databricks credentials not set in environment")
            
            conn = databricks_sql.connect(
                server_hostname=DATABRICKS_HOST.replace("https://", ""),
                http_path=DATABRICKS_HTTP_PATH,
                access_token=DATABRICKS_TOKEN
            )
            logger.info("Connected to Databricks")
            print("✅ Connected to Databricks\n")
        except Exception as e:
            error_msg = f"Failed to connect to Databricks: {str(e)}"
            logger.error(error_msg)
            print(f"❌ {error_msg}")
            return
        
        # Execute query with timeout in Databricks
        print(f"🚀 Executing query in Databricks (timeout: {timeout_minutes} min)...")
        logger.info("Starting Databricks query execution")
        
        query_id, was_killed, error_message, query_metrics = execute_databricks_query_with_timeout(conn, query, timeout_minutes)
        
        # Log query metrics if available
        if query_metrics:
            logger.info("\n" + "="*80)
            logger.info("DATABRICKS QUERY METRICS")
            logger.info("="*80)
            for key, value in query_metrics.items():
                logger.info(f"  {key}: {value}")
            logger.info("="*80)
            
            print("\n" + "="*80)
            print("📊 DATABRICKS QUERY METRICS")
            print("="*80)
            for key, value in query_metrics.items():
                print(f"  {key}: {value}")
            print("="*80 + "\n")
        
        # Close connection
        conn.close()
        
    else:  # Snowflake
        # Connect to Snowflake
        try:
            conn = get_snowflake_connection()
            cursor = conn.cursor()
            logger.info("Connected to Snowflake")
            print("✅ Connected to Snowflake\n")
        except Exception as e:
            error_msg = f"Failed to connect to Snowflake: {str(e)}"
            logger.error(error_msg)
            print(f"❌ {error_msg}")
            return
        
        # Execute query with timeout
        print(f"🚀 Executing query (timeout: {timeout_minutes} min)...")
        logger.info("Starting query execution")
        
        query_id, was_killed, error_message, query_metrics = execute_query_with_timeout(cursor, query, timeout_minutes)
        
        # Log query metrics if available
        if query_metrics:
            logger.info("\n" + "="*80)
            logger.info("SNOWFLAKE QUERY METRICS")
            logger.info("="*80)
            for key, value in query_metrics.items():
                logger.info(f"  {key}: {value}")
            logger.info("="*80)
            
            print("\n" + "="*80)
            print("📊 SNOWFLAKE QUERY METRICS")
            print("="*80)
            for key, value in query_metrics.items():
                print(f"  {key}: {value}")
            print("="*80 + "\n")
    
    # Log and display query ID prominently
    print("\n" + "="*80)
    if query_id:
        logger.info(f"="*80)
        logger.info(f"QUERY ID: {query_id}")
        logger.info(f"="*80)
        print(f"🆔 QUERY ID: {query_id}")
        print("="*80)
    else:
        print("⚠️  Query ID could not be retrieved")
        print("="*80)
    
    # Get query profile using GET_QUERY_OPERATOR_STATS (Snowflake only)
    output_dir = "Output"
    os.makedirs(output_dir, exist_ok=True)
    
    if was_killed or error_message:
        if was_killed:
            print(f"\n⚠️  Query was {('cancelled' if platform.lower() == 'databricks' else 'killed')} after {timeout_minutes} minutes")
            logger.warning(f"Query was {('cancelled' if platform.lower() == 'databricks' else 'killed')} after {timeout_minutes} minutes")
        if error_message:
            print(f"❌ Error: {error_message}")
            logger.error(f"Query error: {error_message}")
    else:
        print("\n✅ Query completed successfully within timeout")
        logger.info("Query completed successfully within timeout")
    
    # Get query profile for all queries (success or failure) - Snowflake only
    if query_id and platform.lower() == 'snowflake':
        print(f"\n📊 Fetching query operator statistics from Snowflake...")
        logger.info(f"Fetching GET_QUERY_OPERATOR_STATS for Query ID: {query_id}")
        
        try:
            operator_stats_query = f"""
            SELECT *
            FROM TABLE(
                GET_QUERY_OPERATOR_STATS('{query_id}')
            )
            """
            
            cursor.execute(operator_stats_query)
            results = cursor.fetchall()
            columns = [desc[0] for desc in cursor.description]
            
            operator_stats_df = pd.DataFrame(results, columns=columns)
            
            if not operator_stats_df.empty:
                # Save to CSV with the specified naming convention
                csv_filename = os.path.join(output_dir, f"{base_filename}_Query_profile_timeline.csv")
                operator_stats_df.to_csv(csv_filename, index=False)
                logger.info(f"Query operator stats saved to: {csv_filename}")
                print(f"✅ Query operator stats saved to: {csv_filename}")
                
                # Display summary
                print(f"\n📊 Query Operator Stats Summary:")
                print(f"   - Total operators: {len(operator_stats_df)}")
                print(f"   - Columns: {len(operator_stats_df.columns)}")
                if 'OPERATOR_TYPE' in operator_stats_df.columns:
                    print(f"   - Operator types: {operator_stats_df['OPERATOR_TYPE'].nunique()}")
            else:
                logger.warning("No operator stats returned from GET_QUERY_OPERATOR_STATS")
                print(f"⚠️  No operator stats returned")
                
        except Exception as e:
            error_msg = f"Failed to fetch query operator stats: {str(e)}"
            logger.error(error_msg)
            print(f"❌ {error_msg}")
    
    elif query_id and platform.lower() == 'databricks':
        print(f"\n⏭️  Skipping query profile for Databricks (not implemented yet)")
        logger.info("Query profile skipped for Databricks platform")
    
    # Close connections (platform-specific)
    if platform.lower() == 'snowflake':
        cursor.close()
        conn.close()
    # Databricks connection already closed
    logger.info("Connections closed")
    logger.info("="*80)
    logger.info("Query Execution and Profiling Completed")
    logger.info("="*80)
    
    print("\n" + "="*80)
    print("✅ QUERY EXECUTION AND PROFILING COMPLETED")
    print("="*80)
    if query_id:
        print(f"🆔 QUERY ID: {query_id}")
        print(f"   (Also logged in: {log_filename})")
    print(f"📄 Log file: {log_filename}")
    print("="*80)
    print("🔧 Build: LOKV-2026")
    print("="*80)
    
    # Final prominent QueryID log entry
    if query_id:
        logger.info("="*80)
        logger.info(f"FINAL QUERY ID: {query_id}")
        logger.info("="*80)
    print()
    
    # Run analysis script automatically based on platform
    if query_id:
        print("\n" + "="*80)
        print("🔍 RUNNING QUERY PERFORMANCE ANALYSIS")
        print("="*80)
        
        if platform.lower() == 'snowflake':
            # Run Snowflake analysis
            analysis_script = "analyze_expensive_nodes.py"
            if os.path.exists(analysis_script):
                print(f"\n📊 Running Snowflake query analysis using {analysis_script}...")
                logger.info(f"Running Snowflake analysis: {analysis_script} --base {base_filename}")
                
                import subprocess
                try:
                    result = subprocess.run(
                        [sys.executable, analysis_script, "--base", base_filename, "--output-dir", output_dir],
                        capture_output=True,
                        text=True,
                        timeout=60
                    )
                    
                    if result.returncode == 0:
                        print(result.stdout)
                        logger.info("Snowflake analysis completed successfully")
                    else:
                        print(f"⚠️  Analysis script completed with warnings:")
                        print(result.stdout)
                        if result.stderr:
                            print(result.stderr)
                        logger.warning(f"Analysis script exit code: {result.returncode}")
                        
                except subprocess.TimeoutExpired:
                    print(f"⚠️  Analysis script timed out after 60 seconds")
                    logger.warning("Analysis script timed out")
                except Exception as e:
                    print(f"⚠️  Could not run analysis script: {e}")
                    logger.warning(f"Failed to run analysis script: {e}")
            else:
                print(f"⚠️  Analysis script not found: {analysis_script}")
                logger.warning(f"Analysis script not found: {analysis_script}")
        
        print("="*80)
    
    # Return query_id for use in other scripts
    return query_id


def get_object_references(
    cursor,
    database: str,
    schema: str,
    object_name: str
) -> List[Dict]:
    """
    Calls Snowflake's GET_OBJECT_REFERENCES to get direct dependencies.
    
    Args:
        cursor: Active Snowflake cursor
        database: Database name
        schema: Schema name
        object_name: Object (view/table) name
        
    Returns:
        List of dictionaries containing dependency information
    """
    query = f"""
        SELECT 
            referenced_database_name,
            referenced_schema_name,
            referenced_object_name,
            referenced_object_type

        FROM TABLE(
            {database}.INFORMATION_SCHEMA.GET_OBJECT_REFERENCES(
                DATABASE_NAME => '{database}',
                SCHEMA_NAME => '{schema}',
                OBJECT_NAME => '{object_name}'
            )
        )
        WHERE referenced_object_type in ('VIEW', 'TABLE') 
        
        
    """
    
    try:
        cursor.execute(query)
        results = cursor.fetchall()
        print('get_reference_object output')
        print(results)
        dependencies = []
        for row in results:
            dependencies.append({
                'referenced_database_name': row[0],
                'referenced_schema_name': row[1],
                'referenced_object_name': row[2],
                'referenced_object_type': row[3]
            })
        
        return dependencies
        # return results
        
    
    except Exception as e:
        print(f"Warning: Could not get references for {database}.{schema}.{object_name}: {str(e)}")
        return []


# conn = get_snowflake_connection()
# cursor = conn.cursor()
# print("✅ Connected to Snowflake (single connection)\n")

def get_clustering_keys(
    cursor,
    database: str,
    schema: str,
    table_name: str
) -> str:
    """
    Gets clustering key information for a table.
    
    Args:
        database: Database name
        schema: Schema name
        table_name: Table name
        
    Returns:
        Comma-separated string of clustering keys, or empty string if none
        Returns "EXCLUDED" for schemas in EXCLUDED_SCHEMAS
    """
    # Check if schema is in excluded list
    if schema.upper() in EXCLUDED_SCHEMAS:
        return "EXCLUDED"
    
    try:
        # conn = get_snowflake_connection()
        # cursor = conn.cursor()
        
        query = f"""
            SELECT CLUSTERING_KEY
            FROM INFORMATION_SCHEMA.TABLES
            where TABLE_NAME = '{table_name}' and TABLE_CATALOG = '{database}'
            AND CLUSTERING_KEY IS NOT NULL
        """
        
        cursor.execute(query)
        result = cursor.fetchone()
        # cursor.close()
        # conn.close()
        
        if result and result[0]:
            # Parse the clustering key (format: LINEAR(col1, col2))
            clustering_key = result[0]
            # Remove LINEAR() wrapper if present
            if clustering_key.startswith('LINEAR(') and clustering_key.endswith(')'):
                clustering_key = clustering_key[7:-1]
            return clustering_key
        return ""
        
    except Exception as e:
        print(f"Warning: Could not get clustering keys for {database}.{table_name}: {str(e)}")
        return "No Cluster Key"
def get_clutser_object_references(
    cursor,
    database: str,
    schema: str,
    object_name: str
) -> Dict:
    """
    Retrieves Snowflake clustering information for the specified object
    using SYSTEM$CLUSTERING_INFORMATION and returns the parsed metrics as a
    dictionary containing total_partition_count, average_overlaps, and average_depth.

    Args:
        cursor: Active Snowflake cursor
        database: Database name (used explicitly in the query)
        schema: Schema name
        object_name: Object (table/view) name

    Returns:
        Dict: {'total_partition_count': int, 'average_overlaps': float, 'average_depth': float}
              Returns an empty dict if no data is available or on error.
    """
    query = f"""
        SELECT SYSTEM$CLUSTERING_INFORMATION('{database}.{schema}.{object_name}') AS INFO;
    """

    try:
        cursor.execute(query)
        result = cursor.fetchall()

        cluster_depth: Dict = {}

        if len(result) < 1:
            print("no cluster information fetched")
        else:
            print("cluster information fetched")

            col_name = f"{object_name}_Information"

            df = pd.DataFrame(result, columns=[col_name])

            df[col_name] = df[col_name].apply(json.loads)

            df['total_partition_count'] = df[col_name].apply(lambda x: x.get('total_partition_count'))
            df['average_overlaps'] = df[col_name].apply(lambda x: x.get('average_overlaps'))
            df['average_depth'] = df[col_name].apply(lambda x: x.get('average_depth'))

            # Safely extract first row values if available
            if not df.empty:
                try:
                    cluster_depth.update([
                        ('total_partition_count', int(df['total_partition_count'].iloc[0]) if pd.notna(df['total_partition_count'].iloc[0]) else 0),
                        ('average_depth', float(df['average_depth'].iloc[0]) if pd.notna(df['average_depth'].iloc[0]) else 0.0),
                        ('average_overlaps', float(df['average_overlaps'].iloc[0]) if pd.notna(df['average_overlaps'].iloc[0]) else 0.0)
                    ])
                except Exception:
                    # If parsing fails for any reason, leave as empty dict
                    pass

        return cluster_depth

    except Exception as e:
        print(f"Warning: Could not get clustering info for {database}.{schema}.{object_name}: {str(e)}")
        return {}


def get_cluster_information_per_clusterKey(
    cursor,
    database: str,
    schema: str,
    object_name: str,
    cluster_keys: Union[str, List[str]]
) -> Dict:
    """
    Retrieves Snowflake clustering metrics for specific clustering key columns
    using SYSTEM$CLUSTERING_INFORMATION with the column list.

    Args:
        cursor: Active Snowflake cursor
        database: Database name
        schema: Schema name
        object_name: Table name
        cluster_keys: A single column name or a list of column names to evaluate

    Returns:
        Dict: {'total_partition_count': int, 'average_overlaps': float, 'average_depth': float}
              Returns an empty dict if no data is available or on error.
    """
    # Normalize to list
    keys: List[str] = [cluster_keys] if isinstance(cluster_keys, str) else list(cluster_keys)
    # Quote each column name for the Snowflake function call
    keys_sql = ", ".join([f"'{k}'" for k in keys])

    query = f"""
        SELECT SYSTEM$CLUSTERING_INFORMATION('{database}.{schema}.{object_name}', ({keys_sql})) AS INFO;
    """

    try:
        cursor.execute(query)
        result = cursor.fetchall()

        cluster_depth: Dict = {}

        if len(result) < 1:
            print("no per-key cluster information fetched")
        else:
            print("per-key cluster information fetched")

            df = pd.DataFrame(result, columns=['INFO'])
            df['INFO'] = df['INFO'].apply(json.loads)

            df['total_partition_count'] = df['INFO'].apply(lambda x: x.get('total_partition_count'))
            df['average_overlaps'] = df['INFO'].apply(lambda x: x.get('average_overlaps'))
            df['average_depth'] = df['INFO'].apply(lambda x: x.get('average_depth'))

            if not df.empty:
                try:
                    cluster_depth.update([
                        ('total_partition_count', int(df['total_partition_count'].iloc[0]) if pd.notna(df['total_partition_count'].iloc[0]) else 0),
                        ('average_depth', float(df['average_depth'].iloc[0]) if pd.notna(df['average_depth'].iloc[0]) else 0.0),
                        ('average_overlaps', float(df['average_overlaps'].iloc[0]) if pd.notna(df['average_overlaps'].iloc[0]) else 0.0)
                    ])
                except Exception:
                    pass

        return cluster_depth

    except Exception as e:
        print(f"Warning: Could not get per-key clustering info for {database}.{schema}.{object_name}: {str(e)}")
        return {}


def _compute_cluster_key_depth(
    cursor,
    database: str,
    schema: str,
    object_name: str,
    clustering_keys_str: str
) -> Dict[str, Dict[str, float]]:
    """
    Compute per-key clustering metrics using SYSTEM$CLUSTERING_INFORMATION for each key.

    Returns a dict mapping key -> { 'average_depth': float, 'average_overlaps': float }.
    Empty dict if keys missing or errors.
    """
    key_depth: Dict[str, Dict[str, float]] = {}
    if not clustering_keys_str:
        return key_depth

    # Parse comma-separated keys and trim whitespace
    keys = [k.strip() for k in clustering_keys_str.split(',') if k and k.strip()]
    for k in keys:
        try:
            metrics = get_cluster_information_per_clusterKey(cursor, database, schema, object_name, k)
            if metrics:
                avg_depth = float(metrics.get('average_depth', 0.0))
                avg_overlaps = float(metrics.get('average_overlaps', 0.0))
                key_depth[k] = {
                    'average_depth': avg_depth,
                    'average_overlaps': avg_overlaps
                }
        except Exception as e:
            print(f"Warning: Could not compute per-key metrics for {database}.{schema}.{object_name} key {k}: {str(e)}")
            # continue to next key
            continue

    return key_depth



def get_databricks_info_for_sf_object(sf_schema: str, sf_object: str) -> Dict[str, str]:
    """
    Get Databricks clustering information for a Snowflake object by mapping SF schema/object to Databricks table name.
    
    Args:
        sf_schema: Snowflake schema name
        sf_object: Snowflake object name
        
    Returns:
        Dict with keys: clusteringColumns, numFiles, name (returns "NONE" for each if not found)
    """
    try:
        # Build Databricks table name: SF_Schema.SF_Object (e.g., "UHCCD.ADJD_MCE")
        databricks_table_name = f"{sf_schema}.{sf_object}"
        
        print(f"   🔍 Getting Databricks info for: {databricks_table_name}")
        
        # Call the Databricks function
        clustering_info = get_databricks_clustering_info(databricks_conn, databricks_table_name)
        
        # Extract the values with additional safeguards for empty arrays
        clustering_columns = clustering_info.get('clusteringColumns', 'NONE')
        
        # Convert lists to comma-separated strings or "NONE"
        if clustering_columns is None or \
           (isinstance(clustering_columns, list) and len(clustering_columns) == 0) or \
           str(clustering_columns).strip() == '' or \
           str(clustering_columns).strip() == '[]':
            clustering_columns = 'NONE'
        elif isinstance(clustering_columns, list):
            # Convert list to comma-separated string
            clustering_columns = ', '.join(str(col) for col in clustering_columns)
        else:
            # Ensure it's a string
            clustering_columns = str(clustering_columns)
            
        num_files = str(clustering_info.get('numFiles', 'NONE'))
        name = str(clustering_info.get('name', 'NONE'))
        size_gb = clustering_info.get('sizeInGB', 'NONE')
        # Convert to string if it's a number
        if isinstance(size_gb, (int, float)):
            size_gb = str(size_gb)
        
        result = {
            'clusteringColumns': clustering_columns,
            'numFiles': num_files,
            'name': name,
            'sizeInGB': size_gb
        }
        
        print(f"   ✅ Databricks info retrieved: {result}")
        return result
        
    except Exception as e:
        print(f"   ❌ Error getting Databricks info for {sf_schema}.{sf_object}: {str(e)}")
        return {
            'clusteringColumns': 'NONE',
            'numFiles': 'NONE', 
            'name': 'NONE',
            'sizeInGB': 'NONE'
        }


# def discover_lineage_recursive(
#     cursor,
#     database: str,
#     schema: str,
#     object_name: str,
#     level: int = 0,
#     parent_object: str = None,
#     visited: Set[Tuple[str, str, str]] = None,
#     all_results: List[Dict] = None
# ) -> List[Dict]:
#     """
#     Recursively discovers object dependencies until base tables are reached.
    
#     Args:
#         database: Database name
#         schema: Schema name
#         object_name: Object name to analyze
#         level: Current recursion depth (0 = starting object)
#         parent_object: Fully qualified name of parent object
#         visited: Set of already processed objects (prevents circular dependencies)
#         all_results: Accumulated results list

#         level 0 : Direct references in the specified object view 
#         level 1 : Refereneces inside level 0 views
        
#     Returns:
#         List of dictionaries with complete lineage information
#     """ 
#     # Initialize collections on first call
#     if visited is None:
#         visited = set()
#     if all_results is None:
#         all_results = []
    
#     # Create unique identifier for current object
#     current_object_key = (database.upper(), schema.upper(), object_name.upper())
#     current_object_fqn = f"{database}.{schema}.{object_name}"
    
#     # Check for circular dependency
#     if current_object_key in visited:
#         print(f"  {'  ' * level}⚠️  Circular dependency detected: {current_object_fqn} (skipping)")
#         return all_results
    
#     # Mark as visited
#     visited.add(current_object_key)
    
#     # Log current processing
#     indent = "  " * level
#     print(f"{indent}🔍 Level {level}: Analyzing {current_object_fqn}")
    
#     # Get direct dependencies
#     dependencies = get_object_references(cursor, database, schema, object_name)
    
#     if not dependencies:
#         print(f"{indent}  ✓ No dependencies found (base table or empty)")
#         return all_results
    
#     # Process each dependency
#     for dep in dependencies:
#         ref_db = dep['referenced_database_name']
#         ref_schema = dep['referenced_schema_name']
#         ref_object = dep['referenced_object_name']
#         ref_type = dep['referenced_object_type']
#         ref_fqn = f"{ref_db}.{ref_schema}.{ref_object}"
        

#         # Get clustering key information if it's a TABLE (not VIEW)
#         clustering_keys = ""
#         if ref_type == 'TABLE':
#             clustering_keys = get_clustering_keys(cursor, ref_db, ref_object)
#             if clustering_keys:
#                 print(f"{indent}  🔑 Clustering keys found: {clustering_keys}")

#         # Add to results
#         result_entry = {
#             'level': level,
#             'parent_object': parent_object if parent_object else current_object_fqn,
#             'current_object': current_object_fqn,
#             'referenced_database': ref_db,
#             'referenced_schema': ref_schema,
#             'referenced_object': ref_object,
#             'referenced_object_type': ref_type,
#             'referenced_object_fqn': ref_fqn,
#             'clustering_keys': clustering_keys
#         }
#         all_results.append(result_entry)
        
#         # Log dependency
#         icon = "📊" if ref_type == "VIEW" else "📋"
#         print(f"{indent}  {icon} Found {ref_type}: {ref_fqn}")
        
#         # If it's a VIEW, recurse into it
#         if ref_type == 'VIEW':
#             print(f'{indent}    ➡️  Recursing into VIEW: {ref_fqn}')
#             discover_lineage_recursive(
#                 cursor=cursor,
#                 database=ref_db,
#                 schema=ref_schema,
#                 object_name=ref_object,
#                 level=level + 1,
#                 parent_object=current_object_fqn,
#                 visited=visited,
#                 all_results=all_results
#             )
#         else:
#             # Base table reached
#             print(f"{indent}    ✓ Base table reached")
    
#     return all_results


def discover_lineage_recursive(
    cursor,
    database: str,
    schema: str,
    object_name: str,
    level: int = 0,
    parent_object: str = None,
    visited: Set[Tuple[str, str, str]] = None,
    all_results: List[Dict] = None
) -> List[Dict]:
    """
    Recursively discovers object dependencies until base tables are reached.
    
    Args:
        database: Database name
        schema: Schema name
        object_name: Object name to analyze
        level: Current recursion depth (0 = starting object)
        parent_object: Fully qualified name of parent object
        visited: Set of already processed objects (prevents circular dependencies)
        all_results: Accumulated results list

        level 0 : Direct references in the specified object view 
        level 1 : Refereneces inside level 0 views
        
    Returns:
        List of dictionaries with complete lineage information
    """ 
    # Initialize collections on first call
    if visited is None:
        visited = set()
    if all_results is None:
        all_results = []
    
    # Create unique identifier for current object
    current_object_key = (database.upper(), schema.upper(), object_name.upper())
    current_object_fqn = f"{database}.{schema}.{object_name}"
    
    # Check for circular dependency
    if current_object_key in visited:
        print(f"  {'  ' * level}⚠️  Circular dependency detected: {current_object_fqn} (skipping)")
        return all_results
    
    # Mark as visited
    visited.add(current_object_key)
    
    # Log current processing
    indent = "  " * level
    print(f"{indent}🔍 Level {level}: Analyzing {current_object_fqn}")
    
    # Get direct dependencies
    dependencies = get_object_references(cursor, database, schema, object_name)
    # De-duplicate dependencies for this object (some views may list the same TABLE multiple times)
    if dependencies:
        unique_deps = []
        seen_keys = set()
        for d in dependencies:
            key = (
                d.get('referenced_database_name'),
                d.get('referenced_schema_name'),
                d.get('referenced_object_name'),
                d.get('referenced_object_type'),
            )
            if key not in seen_keys:
                seen_keys.add(key)
                unique_deps.append(d)
        dependencies = unique_deps
    
    if not dependencies:
        print(f"{indent}  ✓ No dependencies found (base table or empty)")
        return all_results
    
    # Process each dependency
    for dep in dependencies:
        ref_db = dep['referenced_database_name']
        ref_schema = dep['referenced_schema_name']
        ref_object = dep['referenced_object_name']
        ref_type = dep['referenced_object_type']
        ref_fqn = f"{ref_db}.{ref_schema}.{ref_object}"
        

        # Get clustering key information if it's a TABLE (not VIEW)
        clustering_keys = ""
        cluster_depth = {}
        cluster_key_depth = {}
        if ref_type == 'TABLE':
            clustering_keys = get_clustering_keys(cursor, ref_db, ref_schema, ref_object)
            if clustering_keys and clustering_keys not in ["EXCLUDED", ""]:
                print(f"{indent}  🔑 Clustering keys found: {clustering_keys}")
            # Fetch clustering metrics for this table (skip if schema is excluded)
            if ref_schema.upper() not in EXCLUDED_SCHEMAS:
                cluster_depth = get_clutser_object_references(cursor, ref_db, ref_schema, ref_object)
                if cluster_depth:
                    print(f"{indent}  📐 Cluster metrics: {cluster_depth}")
                # Compute per-key average_depth
                if clustering_keys and clustering_keys != "EXCLUDED":
                    cluster_key_depth = _compute_cluster_key_depth(cursor, ref_db, ref_schema, ref_object, clustering_keys)
                    if cluster_key_depth:
                        print(f"{indent}  📏 Per-key average depth: {cluster_key_depth}")

        # Add to results
        # Build combined cluster metric dictionary
        cluster_metric = {
            'clustering_information': cluster_depth,
            'cluster_key_Level_information': cluster_key_depth
        }

        result_entry = {
            'level': level,
            'parent_object': parent_object if parent_object else current_object_fqn,
            'current_object': current_object_fqn,
            'referenced_database': ref_db,
            'referenced_schema': ref_schema,
            'referenced_object': ref_object,
            'referenced_object_type': ref_type,
            'referenced_object_fqn': ref_fqn,
            'clustering_keys': clustering_keys,
            'cluster_depth': cluster_depth,
            'cluster_key_depth': cluster_key_depth,
            'cluster_metric': cluster_metric
        }
        all_results.append(result_entry)
        
        # Log dependency
        icon = "📊" if ref_type == "VIEW" else "📋"
        print(f"{indent}  {icon} Found {ref_type}: {ref_fqn}")
        
        # If it's a VIEW, recurse into it
        if ref_type == 'VIEW':
            print(f'{indent}    ➡️  Recursing into VIEW: {ref_fqn}')
            discover_lineage_recursive(
                cursor=cursor,
                database=ref_db,
                schema=ref_schema,
                object_name=ref_object,
                level=level + 1,
                parent_object=current_object_fqn,
                visited=visited,
                all_results=all_results
            )
        else:
            # Base table reached
            print(f"{indent}    ✓ Base table reached")
    
    return all_results


    
def save_lineage_to_csv(results: List[Dict], view_name: str):
    """
    Saves lineage results to a CSV file named after the view.
    
    This function exports the complete dependency tree discovered during lineage analysis.
    Each row in the CSV represents a dependency relationship with its depth level.
    
    The CSV includes a 'dependency_depth' column with descriptive labels:
    - "DIRECT" for Level 0 (immediate dependencies)
    - "NESTED-1" for Level 1 (dependencies of direct dependencies)
    - "NESTED-2" for Level 2 (and so on...)
    
    DEPENDENCY DEPTH EXPLANATION:
    -----------------------------
    The 'level' column in the CSV represents the dependency depth from your target view.
    To make it clearer, we use descriptive labels instead of just numbers:
    
    Level 0 = "DIRECT" (Direct Dependencies):
        - These are the immediate references found in your target view
        - Example: If analyzing view "ADJD_MCE", DIRECT level shows all tables and views 
          that "ADJD_MCE" directly references in its SQL definition
        - Question answered: "What does my target view directly depend on?"
        
    Level 1 = "NESTED-1" (First-Level Nested Dependencies):
        - These are the dependencies of your DIRECT level views
        - If a DIRECT object is a VIEW (not a base table), we look into that view 
          to find what IT depends on
        - Question answered: "What do my direct dependencies depend on?"
        
    Level 2 = "NESTED-2" (Second-Level Nested Dependencies):
        - These are the dependencies of your NESTED-1 level views
        - The pattern continues for each VIEW found at previous levels
        - Recursion stops when base TABLEs are reached (tables have no dependencies)
        
    Level 3+ = "NESTED-3", "NESTED-4", etc. (Deeper Nested Dependencies)
        
    Example Hierarchy:
        DIRECT (Level 0): ADJD_MCE (target view)
                           ├─→ USER_SECUR_KEY_SK (VIEW - direct reference)
                           ├─→ TABLE_B (TABLE - direct reference, stops here)
                           └─→ VIEW_C (VIEW - direct reference)
        
        NESTED-1 (Level 1): Dependencies inside DIRECT level views
                             USER_SECUR_KEY_SK depends on:
                             ├─→ TABLE_X (stops here)
                             └─→ VIEW_Y (continues to NESTED-2)
                             
                             VIEW_C depends on:
                             └─→ TABLE_Z (stops here)
        
        NESTED-2 (Level 2): Dependencies inside NESTED-1 level views
                             VIEW_Y depends on:
                             └─→ TABLE_W (stops here)
    
    Args:
        results: List of lineage dictionaries containing dependency information
        view_name: Name of the view (used for CSV filename)
    """
    import csv
    
    # Create filename from view name
    filename = f"{view_name}_lineage.csv"
    
    # Add descriptive dependency depth labels to make CSV more readable
    # Convert numeric levels to descriptive labels: 0 -> "DIRECT", 1 -> "NESTED-1", etc.
    for result in results:
        level = result['level']
        if level == 0:
            result['dependency_depth'] = "DIRECT"
        else:
            result['dependency_depth'] = f"NESTED-{level}"
    
    # Define CSV columns with dependency_depth as the first column for clarity
    fieldnames = [
        'dependency_depth',
        'level',
        'parent_object',
        'current_object',
        'referenced_database',
        'referenced_schema',
        'referenced_object',
        'referenced_object_type',
        'referenced_object_fqn',
        'clustering_keys'
    ]
    
    try:
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(results)
        
        print(f"\n💾 Lineage saved to CSV: {filename}")
        print(f"   📊 Total rows: {len(results)}")
        
    except Exception as e:
        print(f"\n❌ Error saving CSV: {str(e)}")

def main(TARGET_DATABASE: str, TARGET_SCHEMA: str, TARGET_OBJECT: str, conn=None, cursor=None):
    """
    Main execution function - discovers lineage and creates an Excel file with separate tabs per view.
    
    Args:
        TARGET_DATABASE: Snowflake database name
        TARGET_SCHEMA: Snowflake schema name
        TARGET_OBJECT: Object (view/table) name
        conn: Optional existing Snowflake connection (if None, creates new one)
        cursor: Optional existing Snowflake cursor (if None, creates new one)
    """
    import pandas as pd
    from datetime import datetime
    global logger
    
    # Track if we created the connection (so we know whether to close it)
    created_connection = False
    
    # Values provided via function arguments (from command-line)
    # Create output directories
    output_dir = "Output"
    logs_dir = "logs"
    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(logs_dir, exist_ok=True)
    
    # Setup logging with TARGET_OBJECT name
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_filename = os.path.join(logs_dir, f"{TARGET_OBJECT}_{timestamp}.log")
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_filename, encoding='utf-8'),
            logging.StreamHandler()  # Also print to console
        ],
        force=True  # Override any existing configuration
    )
    logger = logging.getLogger(__name__)
    
    logger.info("="*80)
    logger.info("SNOWFLAKE LINEAGE DISCOVERY TOOL - Starting")
    logger.info("="*80)
    
    # Teradata configuration
    TERADATA_SCHEMA = "UDWBASESECUREVIEW"
    TERADATA_VIEW = TARGET_OBJECT  # Use same view name as Snowflake
    
    # Schema mapping: Snowflake schema -> Teradata schema
    SCHEMA_MAPPING = {
        "UHCCDSECURITY": "UDWSECURITY",
        "UHCCDSECURITYVIEW": "UDWSECURITY",
        "UDWBASESECUREVIEW1": "UDWBASE",
        "UDWBASESECUREVIEW2": "UDWBASE",
        "UDWBASESECUREVIEW3": "UDWBASE",
        "UHCCDBASEVIEW1" : "UDWBASE",
        "UHCCD" : "UDWBASE",
        "GALAXY" : "GALAXY",
        "UAHDMSECUREVIEW1" : "UDWDM",
        "UAHDMSECUREVIEW2" : "UDWDM",
        "UAHDMSECUREVIEW3" : "UDWDM",
        "UAHDMVIEW1" : "UDWDM",
        "UAHDM" : "UDWDM",
        "UAHSECURITY" : "UDWSECURITYVIEW"

        # Add more mappings as needed
    }
    
    def map_schema_to_teradata(snowflake_schema: str) -> str:
        """Map Snowflake schema to Teradata schema, return original if no mapping exists."""
        mapped = SCHEMA_MAPPING.get(snowflake_schema.upper(), snowflake_schema)
        if mapped != snowflake_schema:
            logger.info(f"Schema mapping: {snowflake_schema} -> {mapped}")
        return mapped

    logger.info(f"Target: {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_OBJECT}")
    logger.info(f"Teradata Schema: {TERADATA_SCHEMA}")
    logger.info(f"Schema mappings configured: {SCHEMA_MAPPING}")
    
    # Use provided connection or create new one
    if conn is None or cursor is None:
        conn = get_snowflake_connection()
        cursor = conn.cursor()
        created_connection = True
        print("✅ Connected to Snowflake (new connection)\n")
        logger.info("Connected to Snowflake successfully (new connection)")
    else:
        print("✅ Reusing existing Snowflake connection\n")
        logger.info("Reusing existing Snowflake connection")

    print("="*80) 
    print("🚀 SNOWFLAKE LINEAGE DISCOVERY TOOL (Excel Output with Tabs per View)")
    print("="*80)
    print(f"\n📍 Target Object: {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_OBJECT}")
    print(f"⏳ Discovering complete lineage tree...\n")

    # Step 1: Get the complete lineage tree for the target view
    print("STEP 1: Discovering all dependencies recursively...\n")
    logger.info("STEP 1: Starting lineage discovery")
    
    complete_lineage = discover_lineage_recursive(
        cursor=cursor,
        database=TARGET_DATABASE,
        schema=TARGET_SCHEMA,
        object_name=TARGET_OBJECT,
        level=0
    )
    
    if not complete_lineage:
        print("⚠️  No dependencies found (base table or object has no upstream lineage).")
        logger.warning("No dependencies found for target object")
        if created_connection:
            cursor.close()
            conn.close()
        # Raise exception so batch processor can track this as SKIPPED
        raise ValueError("NO_DEPENDENCIES: Object has no upstream lineage (likely a base table)")
    
    logger.info(f"Lineage discovery completed: {len(complete_lineage)} dependencies found")
    
    # Step 2: Get Teradata partition columns for all TABLE objects in lineage
    print("\n" + "="*80)
    print("STEP 2: Getting Teradata partition columns for TABLE objects in lineage...")
    print("="*80 + "\n")
    logger.info("="*80)
    logger.info("STEP 2: Getting Teradata partition columns for TABLE objects")
    logger.info("="*80)
    
    # Collect all unique TABLE objects from the lineage
    all_teradata_data = []
    unique_tables = set()
    schema_map_tracking = {}  # Track original SF schema to TD schema mapping
    
    # Get ALL unique TABLE objects from the lineage for which TD partition data will be queried
    # NOTE: Do not add the target object unconditionally here because the target is usually a VIEW.
    #       Cluster/partition keys apply to TABLEs only. We'll only include TABLE dependencies below.
    
    # Add referenced TABLE objects from lineage (skip VIEWs and excluded schemas)
    # This ensures we only query Teradata for base tables
    for entry in complete_lineage:
        ref_schema = entry['referenced_schema']
        ref_object = entry['referenced_object']
        ref_type = entry['referenced_object_type']

        # Skip excluded schemas
        if ref_schema.upper() in EXCLUDED_SCHEMAS:
            continue

        if str(ref_type).upper() == 'TABLE':
            td_schema = map_schema_to_teradata(ref_schema)
            unique_tables.add((td_schema, ref_object))
            schema_map_tracking[(td_schema, ref_object)] = ref_schema
    
    logger.info(f"Found {len(unique_tables)} unique TABLE objects to query in Teradata (from lineage)")
    print(f"   Found {len(unique_tables)} unique TABLE objects to query in Teradata (from lineage)\n")
    
    # Get Teradata partition columns for each unique object (TABLE only)
    for i, (schema, obj_name) in enumerate(unique_tables, 1):
        log_msg = f"{i}/{len(unique_tables)}: Fetching Teradata partition columns for {schema}.{obj_name}"
        print(f"   {log_msg}...", end=" ")
        logger.info(log_msg)
        
        try:
            td_df = get_partition_columns(schema, obj_name, output_csv=False)
            
            if td_df is not None and not td_df.empty:
                # Add source tracking columns
                td_df['TD_Schema'] = schema
                td_df['TD_Object'] = obj_name
                td_df['SF_Schema'] = schema_map_tracking.get((schema, obj_name), schema)
                td_df['SF_Object'] = obj_name
                all_teradata_data.append(td_df)
                result_msg = f"✅ {len(td_df)} rows"
                print(result_msg)
                logger.info(f"Success: Retrieved {len(td_df)} partition column rows for {schema}.{obj_name}")
            else:
                # No partition columns found - add a placeholder row
                result_msg = "⚠️ No partition columns found"
                print(result_msg)
                logger.warning(f"No partition columns found for {schema}.{obj_name}")
                
                # Create a placeholder DataFrame with "No partition key specified"
                placeholder_df = pd.DataFrame({
                    'DatabaseName': [schema],
                    'TableName': [obj_name],
                    'ColumnName': ['No partition key specified'],
                    'ColumnFormat': ['N/A'],
                    'PartitioningColumn': ['N/A'],
                    'TD_Schema': [schema],
                    'TD_Object': [obj_name],
                    'SF_Schema': [schema_map_tracking.get((schema, obj_name), schema)],
                    'SF_Object': [obj_name]
                })
                all_teradata_data.append(placeholder_df)
        except Exception as e:
            error_msg = f"❌ Error: {str(e)}"
            print(error_msg)
            logger.error(f"Error fetching partition columns for {schema}.{obj_name}: {str(e)}")
    
    # Combine all Teradata data
    if all_teradata_data:
        teradata_df = pd.concat(all_teradata_data, ignore_index=True)
        
        # Remove duplicates from Teradata data
        td_initial_count = len(teradata_df)
        teradata_df = teradata_df.drop_duplicates()
        td_removed_duplicates = td_initial_count - len(teradata_df)
        
        if td_removed_duplicates > 0:
            print(f"   🔍 Removed {td_removed_duplicates} duplicate rows from Teradata data")
            logger.info(f"Removed {td_removed_duplicates} duplicate Teradata rows (from {td_initial_count} to {len(teradata_df)})")
        
        # Consolidate partition columns: One row per object with comma-separated partition keys
        print(f"\n   📋 Consolidating partition columns by object...")
        logger.info("Consolidating partition columns by object")
        
        teradata_consolidated = teradata_df.groupby(['TD_Schema', 'TD_Object', 'SF_Schema', 'SF_Object']).agg({
            # Preserve first-seen order of partition columns across rows
            'ColumnName': lambda x: ', '.join(list(dict.fromkeys(x)))
        }).reset_index()
        
        # Rename column for clarity
        teradata_consolidated.rename(columns={'ColumnName': 'TD_Partition_Keys'}, inplace=True)
        
        summary_msg = f"Consolidated to {len(teradata_consolidated)} unique objects (from {len(teradata_df)} individual partition columns)"
        print(f"   📊 {summary_msg}")
        logger.info(summary_msg)

        # Filter teradata_consolidated to include only Snowflake TABLE objects present in lineage
        # Build a set of TABLE object names from lineage
        try:
            sf_table_names = {
                str(entry.get('referenced_object')).upper()
                for entry in complete_lineage
                if str(entry.get('referenced_object_type', '')).upper() == 'TABLE'
            }
        except Exception:
            sf_table_names = set()

        if sf_table_names:
            before_cnt = len(teradata_consolidated)
            teradata_consolidated = teradata_consolidated[
                teradata_consolidated['SF_Object'].str.upper().isin(sf_table_names)
            ].reset_index(drop=True)
            after_cnt = len(teradata_consolidated)
            filtered_out = before_cnt - after_cnt
            if filtered_out > 0:
                msg = f"Filtered out {filtered_out} non-table or unmatched objects from Teradata list (keeping only SF TABLEs)"
                print(f"   🧹 {msg}")
                logger.info(msg)
    else:
        teradata_consolidated = None
        warning_msg = "No Teradata partition data collected"
        print(f"\n   ⚠️  {warning_msg}")
        logger.warning(warning_msg)
    
    # Step 4: Get Databricks clustering info for all TABLE objects in lineage
    print("\n" + "="*80)
    print("STEP 4: Getting Databricks clustering info for TABLE objects in lineage...")
    print("="*80 + "\n")
    logger.info("="*80)
    logger.info("STEP 4: Getting Databricks clustering info for TABLE objects")
    logger.info("="*80)
    
    # Collect Databricks data for all unique TABLE objects from the lineage
    all_databricks_data = []
    
    # Get unique TABLE objects from lineage (same logic as Teradata, skip excluded schemas)
    unique_sf_tables = set()
    for entry in complete_lineage:
        ref_schema = entry['referenced_schema']
        ref_object = entry['referenced_object']
        ref_type = entry['referenced_object_type']

        # Skip excluded schemas 
        if ref_schema.upper() in EXCLUDED_SCHEMAS:
            continue

        if str(ref_type).upper() == 'TABLE':
            unique_sf_tables.add((ref_schema, ref_object))
    
    logger.info(f"Found {len(unique_sf_tables)} unique TABLE objects to query in Databricks (from lineage)")
    print(f"   Found {len(unique_sf_tables)} unique TABLE objects to query in Databricks (from lineage)\n")
    
    # Get Databricks clustering info for each unique TABLE object
    for i, (sf_schema, sf_object) in enumerate(unique_sf_tables, 1):
        log_msg = f"{i}/{len(unique_sf_tables)}: Fetching Databricks clustering info for {sf_schema}.{sf_object}"
        print(f"   {log_msg}...", end=" ")
        logger.info(log_msg)
        
        try:
            db_info = get_databricks_info_for_sf_object(sf_schema, sf_object)
            
            # Create a row for this object
            db_row = {
                'SF_Schema': sf_schema,
                'SF_Object': sf_object,
                'DB_clusteringColumns': db_info['clusteringColumns'],
                'DB_numFiles': db_info['numFiles'],
                'DB_name': db_info['name'],
                'DB_sizeInGB': db_info['sizeInGB']
            }
            all_databricks_data.append(db_row)
            
            result_msg = f"✅ Retrieved"
            print(result_msg)
            logger.info(f"Success: Retrieved Databricks info for {sf_schema}.{sf_object}")
            
        except Exception as e:
            error_msg = f"❌ Error: {str(e)}"
            print(error_msg)
            logger.error(f"Error fetching Databricks info for {sf_schema}.{sf_object}: {str(e)}")
            
            # Add a placeholder row with NONE values
            db_row = {
                'SF_Schema': sf_schema,
                'SF_Object': sf_object,
                'DB_clusteringColumns': 'NONE',
                'DB_numFiles': 'NONE',
                'DB_name': 'NONE',
                'DB_sizeInGB': 'NONE'
            }
            all_databricks_data.append(db_row)
    
    # Create Databricks DataFrame
    if all_databricks_data:
        databricks_df = pd.DataFrame(all_databricks_data)
        print(f"   📊 Created Databricks DataFrame with {len(databricks_df)} objects")
        logger.info(f"Created Databricks DataFrame with {len(databricks_df)} objects")
    else:
        databricks_df = None
        warning_msg = "No Databricks clustering data collected"
        print(f"\n   ⚠️  {warning_msg}")
        logger.warning(warning_msg)
    
    # Step 5: Create comparison between Teradata partition keys and Snowflake clustering keys
    print("\n" + "="*80)
    print("STEP 5: Creating comparison between TD partition keys and SF clustering keys...")
    print("="*80 + "\n")
    logger.info("="*80)
    logger.info("STEP 5: Creating TD vs SF comparison")
    logger.info("="*80)
    
    comparison_data = []
    
    if teradata_consolidated is not None:
        # Build a map of objects to their Snowflake clustering keys from the lineage tabs
        # The clustering keys are in the individual tab data (grouped by current_object)
        print(f"   🔍 Extracting Snowflake clustering keys from lineage data...")
        
        sf_clustering_map = {}
        
        # Method 1: Extract from current_object tabs
        # Each tab represents an object, and contains its dependencies with clustering_keys
        df_all_temp = pd.DataFrame(complete_lineage)
        
        # Group by current_object (the object that has the tab)
        for current_obj_fqn, group_df in df_all_temp.groupby('current_object'):
            # Extract object name from FQN (database.schema.object)
            obj_name = current_obj_fqn.split('.')[-1] if '.' in current_obj_fqn else current_obj_fqn
            
            # Get clustering keys from the rows in this tab
            # Look for rows where this object appears as referenced_object
            obj_rows = group_df[group_df['referenced_object'] == obj_name]
            
            if not obj_rows.empty:
                # Get clustering_keys from the first matching row
                clustering = obj_rows.iloc[0]['clustering_keys']
                if clustering and clustering not in ['', 'No Cluster Key', 'EXCLUDED']:
                    sf_clustering_map[obj_name] = clustering
                    logger.info(f"Found clustering keys for {obj_name} from tab data: {clustering}")
        
        # Method 2: Also add target object's clustering keys directly (skip if excluded schema)
        if TARGET_SCHEMA.upper() not in EXCLUDED_SCHEMAS:
            target_clustering = get_clustering_keys(cursor, TARGET_DATABASE, TARGET_SCHEMA, TARGET_OBJECT)
            if target_clustering and target_clustering not in ['', 'No Cluster Key', 'EXCLUDED']:
                sf_clustering_map[TARGET_OBJECT] = target_clustering
                logger.info(f"Target object {TARGET_OBJECT} clustering keys: {target_clustering}")
        
        # Method 3: Add any remaining objects from referenced_object in lineage
        for entry in complete_lineage:
            obj_name = entry['referenced_object']
            if obj_name not in sf_clustering_map:
                clustering_keys = entry.get('clustering_keys', '')
                if clustering_keys and clustering_keys not in ['', 'No Cluster Key', 'EXCLUDED']:
                    sf_clustering_map[obj_name] = clustering_keys
        
        print(f"   📊 Built clustering key map for {len(sf_clustering_map)} objects")
        logger.info(f"Built clustering key map for {len(sf_clustering_map)} objects")
        logger.info(f"Objects with clustering keys: {list(sf_clustering_map.keys())}")
        
        # Compare each Teradata object with Snowflake clustering keys
        # Build metrics map for cluster_metric serialization per SF object
        metrics_map: Dict[str, str] = {}
        for e in complete_lineage:
            try:
                obj = e.get('referenced_object')
                cd = e.get('cluster_depth', {})
                ck = e.get('cluster_key_depth', {})
                serialized = 'NONE'
                if cd or ck:
                    try:
                        serialized = json.dumps({'clustering_depth': cd, 'cluster_key_depth': ck})
                    except Exception:
                        serialized = 'NONE'
                if obj:
                    # Prefer first non-NONE serialized value
                    if obj not in metrics_map or metrics_map[obj] == 'NONE':
                        metrics_map[obj] = serialized
            except Exception:
                # Skip any problematic entries silently
                continue

        for _, row in teradata_consolidated.iterrows():
            td_schema = row['TD_Schema']
            td_object = row['TD_Object']
            sf_object = row['SF_Object']
            td_keys = row['TD_Partition_Keys']
            
            # Get Snowflake clustering keys for this object
            sf_keys = sf_clustering_map.get(sf_object, '')
            
            # Parse keys into sets for comparison (case-insensitive)
            if td_keys and td_keys != 'No partition key specified':
                td_keys_set = set(k.strip().upper() for k in td_keys.split(','))
            else:
                td_keys_set = set()
            
            if sf_keys and sf_keys not in ['', 'No Cluster Key']:
                sf_keys_set = set(k.strip().upper() for k in sf_keys.split(','))
            else:
                sf_keys_set = set()
            
            # Compare keys
            keys_in_both = td_keys_set & sf_keys_set
            keys_only_in_td = td_keys_set - sf_keys_set
            keys_only_in_sf = sf_keys_set - td_keys_set
            
            # Order-aware comparison: build normalized lists
            td_list = [k.strip().upper() for k in td_keys.split(',')] if td_keys and td_keys not in ['', 'No partition key specified'] else []
            sf_list = [k.strip().upper() for k in sf_keys.split(',')] if sf_keys and sf_keys not in ['', 'No Cluster Key'] else []

            if td_list and sf_list:
                if td_list == sf_list:
                    order_status = "Exactly Matched and in Same Order"
                elif set(td_list) == set(sf_list):
                    order_status = "Matched but in different order"
                else:
                    order_status = "N/A"
            else:
                order_status = "N/A"

            # Determine status
            if not td_keys_set and not sf_keys_set:
                status = "No keys in either"
            elif td_keys_set == sf_keys_set:
                status = "Exact match"
            elif keys_in_both:
                status = "Partial match"
            else:
                status = "No common keys"
            
            comparison_data.append({
                'Object_Name': sf_object,
                'TD_Schema': td_schema,
                'SF_Schema': row['SF_Schema'],
                'Status': status,
                'Order_Match_Status': order_status,
                'TD_Partition_Keys': td_keys if td_keys else 'None',
                'SF_Clustering_Keys': sf_keys if sf_keys else 'None',
                'Keys_In_Both': ', '.join(sorted(keys_in_both)) if keys_in_both else 'None',
                'Keys_Only_In_TD': ', '.join(sorted(keys_only_in_td)) if keys_only_in_td else 'None',
                'Keys_Only_In_SF': ', '.join(sorted(keys_only_in_sf)) if keys_only_in_sf else 'None',
                'cluster_metric': metrics_map.get(sf_object, 'NONE')
            })
        
        comparison_df = pd.DataFrame(comparison_data)
        print(f"   📊 Created comparison for {len(comparison_df)} objects")
        logger.info(f"Created comparison for {len(comparison_df)} objects")
        
        # Count statuses
        if not comparison_df.empty:
            status_counts = comparison_df['Status'].value_counts()
            print(f"   📈 Comparison summary:")
            for status, count in status_counts.items():
                print(f"      {status}: {count}")
                logger.info(f"Comparison - {status}: {count}")
    else:
        comparison_df = None
        print(f"   ⚠️  No comparison data (Teradata data unavailable)")
        logger.warning("No comparison data available")
    
    # Create TD_VS_DB comparison
    print("\n   🔗 Creating TD vs Databricks comparison...")
    logger.info("Creating TD vs Databricks comparison")
    
    td_vs_db_data = []
    
    if teradata_consolidated is not None and databricks_df is not None:
        # Merge Teradata and Databricks data on SF_Object
        for _, td_row in teradata_consolidated.iterrows():
            sf_object = td_row['SF_Object']
            td_keys = td_row['TD_Partition_Keys']
            
            # Find corresponding Databricks row
            db_row = databricks_df[databricks_df['SF_Object'] == sf_object]
            
            if not db_row.empty:
                db_clustering = db_row.iloc[0]['DB_clusteringColumns']
                db_num_files = db_row.iloc[0]['DB_numFiles']
                db_name = db_row.iloc[0]['DB_name']
                db_size_gb = db_row.iloc[0]['DB_sizeInGB']
                
                # Convert to string to handle arrays/Series properly
                db_clustering = str(db_clustering) if db_clustering is not None else 'NONE'
                db_num_files = str(db_num_files) if db_num_files is not None else 'NONE'
                db_name = str(db_name) if db_name is not None else 'NONE'
                db_size_gb = str(db_size_gb) if db_size_gb is not None else 'NONE'
                
                # Additional safeguard: ensure empty arrays are converted to "NONE"
                if db_clustering.strip() == '' or \
                   db_clustering.strip() == '[]' or \
                   db_clustering.strip() == 'None':
                    db_clustering = 'NONE'
            else:
                db_clustering = 'NONE'
                db_num_files = 'NONE'
                db_name = 'NONE'
                db_size_gb = 'NONE'
            
            # Parse keys for comparison
            if td_keys and td_keys != 'No partition key specified':
                td_keys_set = set(k.strip().upper() for k in td_keys.split(','))
            else:
                td_keys_set = set()
            
            # Now db_clustering is guaranteed to be a string
            if db_clustering != 'NONE':
                # Handle different Databricks clustering column formats
                if isinstance(db_clustering, str) and db_clustering.startswith('[') and db_clustering.endswith(']'):
                    # Parse array-like string format, e.g., "['col1', 'col2']"
                    try:
                        import ast
                        db_list = ast.literal_eval(db_clustering)
                        db_keys_set = set(k.strip().upper() for k in db_list)
                    except:
                        db_keys_set = set()
                elif isinstance(db_clustering, str) and ',' in db_clustering:
                    # Parse comma-separated format
                    db_keys_set = set(k.strip().upper() for k in db_clustering.split(','))
                else:
                    # Single column or other format
                    db_keys_set = {str(db_clustering).strip().upper()}
            else:
                db_keys_set = set()
            
            # Compare keys
            keys_in_both = td_keys_set & db_keys_set
            keys_only_in_td = td_keys_set - db_keys_set
            keys_only_in_db = db_keys_set - td_keys_set
            
            # Determine status
            if not td_keys_set and not db_keys_set:
                status = "No keys in either"
            elif td_keys_set == db_keys_set:
                status = "Exact match"
            elif keys_in_both:
                status = "Partial match"
            else:
                status = "No common keys"
            
            td_vs_db_data.append({
                'Object_Name': sf_object,
                'TD_Schema': td_row['TD_Schema'],
                'SF_Schema': td_row['SF_Schema'],
                'Status': status,
                'TD_Partition_Keys': td_keys if td_keys else 'None',
                'DB_Clustering_Columns': db_clustering if db_clustering != 'NONE' else 'None',
                'DB_Num_Files': db_num_files if db_num_files != 'NONE' else 'None',
                'DB_Name': db_name if db_name != 'NONE' else 'None',
                'DB_SizeInGB': db_size_gb if db_size_gb != 'NONE' else 'None',
                'Keys_In_Both': ', '.join(sorted(keys_in_both)) if keys_in_both else 'None',
                'Keys_Only_In_TD': ', '.join(sorted(keys_only_in_td)) if keys_only_in_td else 'None',
                'Keys_Only_In_DB': ', '.join(sorted(keys_only_in_db)) if keys_only_in_db else 'None'
            })
        
        td_vs_db_df = pd.DataFrame(td_vs_db_data)
        print(f"   📊 Created TD vs DB comparison for {len(td_vs_db_df)} objects")
        logger.info(f"Created TD vs DB comparison for {len(td_vs_db_df)} objects")
        
        # Count statuses
        if not td_vs_db_df.empty:
            status_counts = td_vs_db_df['Status'].value_counts()
            print(f"   📈 TD vs DB comparison summary:")
            for status, count in status_counts.items():
                print(f"      {status}: {count}")
                logger.info(f"TD vs DB comparison - {status}: {count}")
    else:
        td_vs_db_df = None
        print(f"   ⚠️  No TD vs DB comparison data (missing Teradata or Databricks data)")
        logger.warning("No TD vs DB comparison data available")
    
    # Step 6: Create Excel file with separate tabs per view
    print("\n" + "="*80)
    print("STEP 6: Creating Excel file with tabs separated by view...")
    print("="*80 + "\n")
    logger.info("="*80)
    logger.info("STEP 6: Creating Excel file")
    logger.info("="*80)
    
    excel_filename = os.path.join(output_dir, f"{TARGET_OBJECT}_lineage_{timestamp}.xlsx")
    
    # Add dependency depth labels
    for entry in complete_lineage:
        level = entry['level']
        if level == 0:
            entry['dependency_depth'] = "DIRECT"
        else:
            entry['dependency_depth'] = f"NESTED-{level}"
    
    # Convert to DataFrame
    df_all = pd.DataFrame(complete_lineage)
    
    # Compute a serialized cluster_metric column (JSON string or 'NONE')
    # This safely represents clustering metrics without keeping dict-typed columns in the Excel output
    def _serialize_cluster_metric(row):
        try:
            cd = row.get('cluster_depth', {}) if isinstance(row, dict) else {}
            ck = row.get('cluster_key_depth', {}) if isinstance(row, dict) else {}
        except Exception:
            cd, ck = {}, {}
        # If df_all rows are Series, access via keys
        if not cd and 'cluster_depth' in df_all.columns:
            try:
                cd = row['cluster_depth'] if isinstance(row.get('cluster_depth', {}), dict) else {}
            except Exception:
                cd = {}
        if not ck and 'cluster_key_depth' in df_all.columns:
            try:
                ck = row['cluster_key_depth'] if isinstance(row.get('cluster_key_depth', {}), dict) else {}
            except Exception:
                ck = {}
        if (not cd) and (not ck):
            return 'NONE'
        try:
            return json.dumps({'clustering_information': cd, 'cluster_key_Level_information': ck})
        except Exception:
            return 'NONE'

    df_all['cluster_metric'] = df_all.apply(_serialize_cluster_metric, axis=1)

    # Reorder columns to put important ones first (avoids dict columns before drop_duplicates)
    column_order = [
        'dependency_depth',
        'level',
        'referenced_object', 
        'referenced_object_type',
        'referenced_database',
        'referenced_schema',
        'referenced_object_fqn',
        'clustering_keys',
        'cluster_metric',
        'parent_object',
        'current_object'
    ]
    
    df_all = df_all[column_order]

    # Remove duplicates from the lineage data using only the selected columns
    initial_row_count = len(df_all)
    df_all = df_all.drop_duplicates()
    removed_duplicates = initial_row_count - len(df_all)
    
    if removed_duplicates > 0:
        print(f"   🔍 Removed {removed_duplicates} duplicate rows from lineage data")
        logger.info(f"Removed {removed_duplicates} duplicate rows (from {initial_row_count} to {len(df_all)})")
    
    # Group by current_object (which represents each view in the lineage tree)
    grouped = df_all.groupby('current_object')
    # Aggregate groups by object name to ensure one sheet per object name (avoids duplicate tabs)
    aggregated_by_obj: Dict[str, pd.DataFrame] = {}
    for current_obj_fqn, group_df in grouped:
        obj_name = current_obj_fqn.split('.')[-1] if '.' in current_obj_fqn else current_obj_fqn
        if obj_name not in aggregated_by_obj:
            aggregated_by_obj[obj_name] = group_df.copy()
        else:
            aggregated_by_obj[obj_name] = pd.concat([aggregated_by_obj[obj_name], group_df], ignore_index=True)

    unique_tabs_count = len(aggregated_by_obj)
    
    # Create Excel writer object
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        # Write summary tab with all data
        df_all.to_excel(writer, sheet_name='All_Lineage', index=False)
        print(f"   ✅ Created tab: 'All_Lineage' ({len(df_all)} rows)")
        
        # Create a tab for each unique object name (combined across schemas)
        for obj_name, combined_df in aggregated_by_obj.items():
            # Excel sheet names have a 31 character limit, truncate if needed
            # Also remove invalid characters
            safe_sheet_name = str(obj_name)[:31]  # Use object name only
            safe_sheet_name = safe_sheet_name.replace('[', '').replace(']', '').replace('*', '').replace('?', '').replace('/', '').replace('\\', '')

            # Augment the tab data: include rows where this object appears as a referenced object
            # This ensures tabs contain both the object's own dependencies and any direct references to it (e.g., DIRECT-0)
            try:
                # Combine and de-duplicate rows across schemas for this object name
                tab_df = combined_df.drop_duplicates().copy()
                # Order rows with DIRECT before NESTED to prioritize direct references
                if 'dependency_depth' in tab_df.columns:
                    tab_df['__depth_sort_key'] = tab_df['dependency_depth'].apply(lambda x: 0 if str(x).upper() == 'DIRECT' else 1)
                    tab_df = tab_df.sort_values(['__depth_sort_key', 'level']).drop(columns=['__depth_sort_key'])
                
                # Do not include cluster_metric in individual object tabs
                if 'cluster_metric' in tab_df.columns:
                    tab_df = tab_df.drop(columns=['cluster_metric'])

                tab_df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
                print(f"   ✅ Created tab: '{safe_sheet_name}' ({len(tab_df)} rows)")
                logger.info(f"Created individual tab for object '{safe_sheet_name}' with {len(tab_df)} rows")
            except Exception as e:
                print(f"   ⚠️  Could not create tab for {obj_name}: {str(e)}")
                logger.error(f"Error creating tab for {obj_name}: {str(e)}")
        
        # Create Teradata Partition Columns tab (consolidated - one row per object)
        if teradata_consolidated is not None and not teradata_consolidated.empty:
            teradata_consolidated.to_excel(writer, sheet_name='TD_Partitions', index=False)
            print(f"   ✅ Created tab: 'TD_Partitions' ({len(teradata_consolidated)} objects)")
            logger.info(f"Created TD_Partitions tab with {len(teradata_consolidated)} objects (consolidated)")
        else:
            print(f"   ⚠️  No Teradata partition data available, skipping TD_Partitions tab")
            logger.warning("No Teradata partition data available, skipping TD_Partitions tab")
        
        # Create Comparison tab
        if comparison_df is not None and not comparison_df.empty:
            comparison_df.to_excel(writer, sheet_name='TD_vs_SF_Comparison', index=False)
            
            # Apply color coding to Order_Match_Status column
            from openpyxl.styles import PatternFill
            ws = writer.sheets['TD_vs_SF_Comparison']
            
            # Find the Order_Match_Status column index
            order_col_idx = None
            for idx, col in enumerate(comparison_df.columns, start=1):
                if col == 'Order_Match_Status':
                    order_col_idx = idx
                    break
            
            if order_col_idx:
                # Define colors for different statuses
                # Green for exact match, Yellow for different order, Gray for N/A
                exact_match_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green
                different_order_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Light yellow
                na_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # Light gray
                
                # Apply colors to each cell in the Order_Match_Status column (skip header row)
                for row_idx in range(2, len(comparison_df) + 2):  # Excel rows start at 1, +1 for header
                    cell = ws.cell(row=row_idx, column=order_col_idx)
                    cell_value = str(cell.value).strip()
                    
                    if cell_value == 'Exactly Matched and in Same Order':
                        cell.fill = exact_match_fill
                    elif cell_value == 'Matched but in different order':
                        cell.fill = different_order_fill
                    elif cell_value == 'N/A':
                        cell.fill = na_fill
                
                print(f"   ✅ Created tab: 'TD_vs_SF_Comparison' ({len(comparison_df)} objects) with color-coded Order_Match_Status")
                logger.info(f"Created TD_vs_SF_Comparison tab with {len(comparison_df)} objects and color coding applied")
            else:
                print(f"   ✅ Created tab: 'TD_vs_SF_Comparison' ({len(comparison_df)} objects)")
                logger.info(f"Created TD_vs_SF_Comparison tab with {len(comparison_df)} objects")
        else:
            print(f"   ⚠️  No comparison data available, skipping TD_vs_SF_Comparison tab")
            logger.warning("No comparison data available, skipping TD_vs_SF_Comparison tab")
        
        # Create Databricks data tab
        if databricks_df is not None and not databricks_df.empty:
            databricks_df.to_excel(writer, sheet_name='Databricks_Info', index=False)
            print(f"   ✅ Created tab: 'Databricks_Info' ({len(databricks_df)} objects)")
            logger.info(f"Created Databricks_Info tab with {len(databricks_df)} objects")
        else:
            print(f"   ⚠️  No Databricks data available, skipping Databricks_Info tab")
            logger.warning("No Databricks data available, skipping Databricks_Info tab")
        
        # Create TD vs DB Comparison tab
        if td_vs_db_df is not None and not td_vs_db_df.empty:
            td_vs_db_df.to_excel(writer, sheet_name='TD_vs_DB', index=False)
            
            # Apply color coding to Status column
            from openpyxl.styles import PatternFill
            ws = writer.sheets['TD_vs_DB']
            
            # Find the Status column index
            status_col_idx = None
            for idx, col in enumerate(td_vs_db_df.columns, start=1):
                if col == 'Status':
                    status_col_idx = idx 
                    break
            
            if status_col_idx:
                # Define colors for different statuses
                exact_match_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green
                partial_match_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Light yellow
                no_common_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Light red
                no_keys_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # Light gray
                
                # Apply colors to each cell in the Status column (skip header row)
                for row_idx in range(2, len(td_vs_db_df) + 2):  # Excel rows start at 1, +1 for header
                    cell = ws.cell(row=row_idx, column=status_col_idx)
                    cell_value = str(cell.value).strip()
                    
                    if cell_value == 'Exact match':
                        cell.fill = exact_match_fill
                    elif cell_value == 'Partial match':
                        cell.fill = partial_match_fill
                    elif cell_value == 'No common keys':
                        cell.fill = no_common_fill
                    elif cell_value == 'No keys in either':
                        cell.fill = no_keys_fill
                
                print(f"   ✅ Created tab: 'TD_vs_DB' ({len(td_vs_db_df)} objects) with color-coded Status")
                logger.info(f"Created TD_vs_DB tab with {len(td_vs_db_df)} objects and color coding applied")
            else:
                print(f"   ✅ Created tab: 'TD_vs_DB' ({len(td_vs_db_df)} objects)")
                logger.info(f"Created TD_vs_DB tab with {len(td_vs_db_df)} objects")
        else:
            print(f"   ⚠️  No TD vs DB comparison data available, skipping TD_vs_DB tab")
            logger.warning("No TD vs DB comparison data available, skipping TD_vs_DB tab")
    
    # Summary statistics
    total_records = len(complete_lineage)
    view_count = sum(1 for item in complete_lineage if item['referenced_object_type'] == 'VIEW')
    table_count = sum(1 for item in complete_lineage if item['referenced_object_type'] == 'TABLE')
    unique_views = unique_tabs_count
    excluded_count = sum(1 for item in complete_lineage if item.get('referenced_schema', '').upper() in EXCLUDED_SCHEMAS)
    
    print(f"\n✅ Excel file created: {excel_filename}")
    print(f"📊 Total dependencies: {total_records}")
    print(f"   - VIEWs: {view_count}")
    print(f"   - TABLEs: {table_count}")
    print(f"   - Excluded schemas (security): {excluded_count} objects")
    print(f"   - Excluded schemas: {', '.join(sorted(EXCLUDED_SCHEMAS))}")
    print(f"   - Tabs created: {unique_views + 1} (1 summary + {unique_views} individual views)")
    print(f"📝 Log file created: {log_filename}")
    print("="*80)
    
    logger.info("="*80)
    logger.info(f"Excel file created: {excel_filename}")
    logger.info(f"Total dependencies: {total_records} (VIEWs: {view_count}, TABLEs: {table_count})")
    logger.info(f"Excluded schemas: {excluded_count} objects from {', '.join(sorted(EXCLUDED_SCHEMAS))}")
    logger.info(f"Tabs created: {unique_views + 1}")
    logger.info(f"Log file: {log_filename}")
    logger.info("Process completed successfully")
    logger.info("="*80)
    
    # Close connections only if we created them
    if created_connection:
        cursor.close()
        conn.close()


import sys

def map_database_name(database_name: str) -> str:
    """
    Map source database names to target Snowflake database names.
    
    This function handles the database naming convention mapping:
    - PROD_UHCCD -> ZJX_PRD_UHCCD_DB
    - Add more mappings as needed
    
    Args:
        database_name (str): Original database name from query/file
    
    Returns:
        str: Mapped Snowflake database name
    
    Example:
        >>> map_database_name('PROD_UHCCD')
        'ZJX_PRD_UHCCD_DB'
        >>> map_database_name('ZJX_PRD_UHCCD_DB')
        'ZJX_PRD_UHCCD_DB'
    """
    if not database_name:
        return 'ZJX_PRD_UHCCD_DB'  # Default database
    
    # Database name mapping
    database_mapping = {
        'PROD_UHCCD': 'ZJX_PRD_UHCCD_DB',
        'UHCCD': 'ZJX_PRD_UHCCD_DB',
        'PROD': 'ZJX_PRD_UHCCD_DB',
    }
    
    # Check if mapping exists (case-insensitive)
    db_upper = database_name.upper().strip()
    if db_upper in database_mapping:
        return database_mapping[db_upper]
    
    # If no mapping found, return original name
    # (assumes it's already in correct format like ZJX_PRD_UHCCD_DB)
    return database_name


def extract_temp_table_names(sql_code: str) -> List[str]:
    """
    Extract temporary table names from SQL CREATE TEMPORARY/VOLATILE TABLE statements.
    
    This function identifies temporary tables that should be excluded from lineage analysis
    since they are transient objects created during query execution.
    
    Patterns Detected:
    ==================
    - CREATE TEMPORARY TABLE table_name
    - CREATE OR REPLACE TEMPORARY TABLE table_name
    - CREATE TEMP TABLE table_name
    - CREATE VOLATILE TABLE table_name (Teradata)
    - CREATE MULTISET VOLATILE TABLE table_name (Teradata)
    - CREATE OR REPLACE VOLATILE TABLE table_name
    - Case-insensitive matching
    
    Args:
        sql_code (str): SQL code to parse
    
    Returns:
        List[str]: List of temporary table names found in uppercase
    
    Example:
        >>> sql = '''
        ... CREATE OR REPLACE TEMPORARY TABLE temp_claims AS
        ... SELECT * FROM claims WHERE date > '2024-01-01';
        ... CREATE MULTISET VOLATILE TABLE VT_ADJD_MCE AS
        ... SELECT * FROM source_table;
        ... '''
        >>> temp_tables = extract_temp_table_names(sql)
        >>> print(temp_tables)
        ['TEMP_CLAIMS', 'VT_ADJD_MCE'] 
    """
    # Remove SQL comments first
    sql_code = re.sub(r'--[^\n]*', ' ', sql_code)
    sql_code = re.sub(r'/\*.*?\*/', ' ', sql_code, flags=re.DOTALL)
    
    # Pattern to match CREATE TEMPORARY/TEMP/VOLATILE TABLE statements
    # Matches: CREATE [OR REPLACE] [MULTISET] TEMPORARY|TEMP|VOLATILE TABLE [schema.]table_name
    # MULTISET is optional (Teradata-specific keyword)
    pattern = r'CREATE\s+(?:OR\s+REPLACE\s+)?(?:MULTISET\s+)?(?:TEMPORARY|TEMP|VOLATILE)\s+TABLE\s+(?:[\w.]+\.)?([\w]+)'
    
    matches = re.findall(pattern, sql_code, re.IGNORECASE)
    
    # Return uppercase names for consistent comparison
    return [match.upper() for match in matches]


def extract_cte_names(sql_code: str) -> List[str]:
    """
    Extract CTE (Common Table Expression) names from SQL WITH clauses.
    
    This function identifies CTEs that should be excluded from lineage analysis
    since they are temporary named result sets within a query.
    
    Patterns Detected:
    ==================
    - WITH cte_name AS (SELECT ...)
    - WITH RECURSIVE cte_name AS (SELECT ...)
    - Multiple CTEs: WITH cte1 AS (...), cte2 AS (...)
    - Case-insensitive matching
    
    Args:
        sql_code (str): SQL code to parse
    
    Returns:
        List[str]: List of CTE names found in uppercase
    
    Example:
        >>> sql = '''
        ... WITH monthly_sales AS (
        ...     SELECT * FROM sales WHERE month = 1
        ... )
        ... SELECT * FROM monthly_sales;
        ... '''
        >>> ctes = extract_cte_names(sql)
        >>> print(ctes)
        ['MONTHLY_SALES']
    """
    # Remove SQL comments first
    sql_code = re.sub(r'--[^\n]*', ' ', sql_code)
    sql_code = re.sub(r'/\*.*?\*/', ' ', sql_code, flags=re.DOTALL)
    
    # Pattern to match CTE names in WITH clauses
    # Matches: WITH [RECURSIVE] cte_name AS
    # Also handles multiple CTEs separated by commas
    pattern = r'WITH\s+(?:RECURSIVE\s+)?([\w]+)\s+AS\s*\(|,\s*([\w]+)\s+AS\s*\('
    
    matches = re.findall(pattern, sql_code, re.IGNORECASE)
    
    # Flatten the tuple results and filter out empty strings
    cte_names = [name.upper() for match in matches for name in match if name]
    
    return cte_names


def extract_subquery_aliases(sql_code: str) -> List[str]:
    """
    Extract subquery aliases from SQL (inline views with AS alias).
    
    This function identifies subquery aliases that should be excluded from lineage analysis
    since they are temporary aliases for inline subqueries, not actual database objects.
    
    Patterns Detected:
    ==================
    - ) AS alias_name
    - ) alias_name (without AS keyword)
    - Handles nested subqueries
    - Case-insensitive matching
    
    Args:
        sql_code (str): SQL code to parse
    
    Returns:
        List[str]: List of subquery aliases found in uppercase
    
    Example:
        >>> sql = '''
        ... SELECT * FROM (
        ...     SELECT * FROM users
        ... ) AS u
        ... JOIN (
        ...     SELECT * FROM orders
        ... ) o ON u.id = o.user_id;
        ... '''
        >>> aliases = extract_subquery_aliases(sql)
        >>> print(aliases)
        ['U', 'O']
    """
    # Remove SQL comments first
    sql_code = re.sub(r'--[^\n]*', ' ', sql_code)
    sql_code = re.sub(r'/\*.*?\*/', ' ', sql_code, flags=re.DOTALL)
    
    # Pattern to match subquery aliases
    # Matches: ) AS alias or ) alias
    # Only capture short aliases (typically 1-3 chars) to avoid false positives
    pattern = r'\)\s+(?:AS\s+)?([a-zA-Z_][a-zA-Z0-9_]{0,10})(?:\s+|$|,|\)|;|ON|WHERE|JOIN|INNER|LEFT|RIGHT|FULL)'
    
    matches = re.findall(pattern, sql_code, re.IGNORECASE)
    
    # Filter out SQL keywords that might be captured
    sql_keywords = {'SELECT', 'FROM', 'WHERE', 'JOIN', 'INNER', 'LEFT', 'RIGHT', 'FULL', 
                    'OUTER', 'CROSS', 'ON', 'AND', 'OR', 'NOT', 'IN', 'EXISTS', 'UNION',
                    'INTERSECT', 'EXCEPT', 'ORDER', 'BY', 'GROUP', 'HAVING', 'LIMIT',
                    'OFFSET', 'FETCH', 'WITH', 'AS', 'CASE', 'WHEN', 'THEN', 'ELSE', 'END'}
    
    aliases = [match.upper() for match in matches if match.upper() not in sql_keywords]
    
    return aliases


def parse_sql_for_objects(sql_code: str) -> List[Dict[str, str]]:
    """
    Parse SQL code to extract database, schema, and object names from FROM and JOIN clauses.
    
    This function analyzes SQL code to find all table/view references and extracts their
    fully qualified names (database.schema.object) or partially qualified names.
    
    Patterns Supported:
    ===================
    1. Fully qualified: database.schema.object
    2. Schema qualified: schema.object
    3. Object only: object
    4. Aliases are ignored (e.g., FROM table AS t1)
    
    Search Locations:
    =================
    - FROM clauses
    - JOIN clauses (INNER JOIN, LEFT JOIN, RIGHT JOIN, FULL JOIN, CROSS JOIN)
    - Handles multi-line SQL
    - Case-insensitive matching
    
    Args:
        sql_code (str): SQL code to parse
    
    Returns:
        List[Dict[str, str]]: List of dictionaries with keys:
            - 'database': Database name (may be empty if not specified)
            - 'schema': Schema name (may be empty if not specified)
            - 'object': Object name (table/view name)
            - 'full_name': Full qualified name as found in SQL
    
    Example:
        >>> sql = '''
        ... SELECT * FROM database1.schema1.table1 t1
        ... JOIN schema2.table2 t2 ON t1.id = t2.id
        ... LEFT JOIN table3 t3 ON t2.id = t3.id
        ... '''
        >>> objects = parse_sql_for_objects(sql)
        >>> print(objects)
        [
            {'database': 'database1', 'schema': 'schema1', 'object': 'table1', 'full_name': 'database1.schema1.table1'},
            {'database': '', 'schema': 'schema2', 'object': 'table2', 'full_name': 'schema2.table2'},
            {'database': '', 'schema': '', 'object': 'table3', 'full_name': 'table3'}
        ]
    """
    # Remove SQL comments
    # Remove single-line comments (--)
    sql_code = re.sub(r'--[^\n]*', ' ', sql_code)
    # Remove multi-line comments (/* */)
    sql_code = re.sub(r'/\*.*?\*/', ' ', sql_code, flags=re.DOTALL)
    
    # Remove EXTRACT(...FROM...) patterns to avoid matching FROM keyword in EXTRACT
    # EXTRACT uses FROM keyword but it's not a table reference
    sql_code = re.sub(r'EXTRACT\s*\([^)]+\sFROM\s+[^)]+\)', ' ', sql_code, flags=re.IGNORECASE)
    
    # Convert to uppercase for case-insensitive matching
    sql_upper = sql_code.upper()
    
    # Common SQL functions to exclude (not tables)
    SQL_FUNCTIONS = {
        'CURRENT_DATE', 'CURRENT_TIME', 'CURRENT_TIMESTAMP', 'GETDATE', 'SYSDATE',
        'NOW', 'DATEADD', 'DATEDIFF', 'EXTRACT', 'CAST', 'CONVERT', 'COALESCE',
        'CASE', 'WHEN', 'THEN', 'ELSE', 'END', 'SUBSTR', 'SUBSTRING', 'TRIM',
        'UPPER', 'LOWER', 'CONCAT', 'LENGTH', 'COUNT', 'SUM', 'AVG', 'MIN', 'MAX',
        'ROW_NUMBER', 'RANK', 'DENSE_RANK', 'LEAD', 'LAG', 'FIRST_VALUE', 'LAST_VALUE'
    }
    
    # Pattern to match FROM and JOIN clauses ONLY
    # This pattern uses strict matching to ensure we only capture table references
    # after FROM/JOIN keywords, not column references in SELECT/WHERE/etc.
    
    # Pattern explanation:
    # (?:FROM|JOIN|...) - Match FROM or various JOIN types
    # \s+ - One or more whitespace
    # ([\w.]+) - Capture database.schema.object (word characters and dots)
    # This ensures we're matching table names in FROM/JOIN context only
    pattern = r'(?:FROM|JOIN|INNER\s+JOIN|LEFT\s+JOIN|RIGHT\s+JOIN|FULL\s+JOIN|CROSS\s+JOIN|LEFT\s+OUTER\s+JOIN|RIGHT\s+OUTER\s+JOIN|FULL\s+OUTER\s+JOIN)\s+([\w.]+)'
    
    matches = re.finditer(pattern, sql_upper, re.IGNORECASE)
    
    objects_found = []
    seen = set()  # To track duplicates
    
    for match in matches:
        full_name = match.group(1)
        
        # Skip if already processed
        if full_name in seen:
            continue
        
        # Split by dots to check if it's a SQL function or column reference
        parts = full_name.split('.')
        
        # Skip if more than 3 parts (invalid: max is database.schema.table)
        if len(parts) > 3:
            continue
        
        # Skip SQL functions (e.g., CURRENT_DATE, GETDATE())
        if len(parts) == 1 and parts[0] in SQL_FUNCTIONS:
            continue
        
        # Skip if the last part (object name) is a SQL function
        if parts[-1] in SQL_FUNCTIONS:
            continue
        
        seen.add(full_name)
        
        # Split by dots to extract database, schema, object
        parts = full_name.split('.')
        
        if len(parts) == 3:
            # Fully qualified: database.schema.object
            obj_dict = {
                'database': parts[0],
                'schema': parts[1],
                'object': parts[2],
                'full_name': full_name
            }
        elif len(parts) == 2:
            # Schema qualified: schema.object
            obj_dict = {
                'database': '',
                'schema': parts[0],
                'object': parts[1],
                'full_name': full_name
            }
        elif len(parts) == 1:
            # Object only: object
            obj_dict = {
                'database': '',
                'schema': '',
                'object': parts[0],
                'full_name': full_name
            }
        else:
            # Invalid format, skip
            continue
        
        objects_found.append(obj_dict)
    
    return objects_found


def get_objects_from_txt_file(txt_file_path: str, default_database: str = None, default_schema: str = None) -> Tuple[str, str, List[str]]:
    """
    Read a .txt file from the source folder and extract distinct database, schema, and view names.
    
    The .txt file can contain:
    - SQL queries (will be parsed for object references)
    - Fully qualified object names (database.schema.object)
    - Schema-qualified names (schema.object)
    - Simple object names (one per line)
    - Comments (lines starting with -- or #)
    
    Args:
        txt_file_path (str): Path to the .txt file
        default_database (str): Default database if not found in file
        default_schema (str): Default schema if not found in file
    
    Returns:
        Tuple[str, str, List[str]]: (database, schema, list of distinct object names)
    
    Example:
        >>> db, schema, objects = get_objects_from_txt_file('source/views.txt')
        >>> print(f"Database: {db}, Schema: {schema}, Objects: {objects}")
        Database: ZJX_PRD_UHCCD_DB, Schema: UDWBASESECUREVIEW1, Objects: ['VIEW1', 'VIEW2']
    """
    if not os.path.exists(txt_file_path):
        print(f"❌ Error: File not found: {txt_file_path}")
        return default_database or '', default_schema or '', []
    
    print(f"\n📂 Reading .txt file: {txt_file_path}")
    
    try:
        with open(txt_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        print(f"   ✅ File loaded ({len(content)} characters)")
    except Exception as e:
        print(f"❌ Error reading file: {str(e)}")
        return default_database or '', default_schema or '', []
    
    # Check if content contains Python variable assignment with SQL (e.g., query = """SELECT...""")
    # Support both triple double quotes """ and triple single quotes '''
    # Handle multiple variable assignments in the same file
    triple_quote_pattern = r'^\s*\w+\s*=\s*("""|\'\'\')(.*?)("""|\'\'\')'
    matches = re.finditer(triple_quote_pattern, content, re.DOTALL | re.MULTILINE)
    
    sql_parts = []
    for match in matches:
        sql_content = match.group(2).strip()
        if sql_content:
            sql_parts.append(sql_content)
    
    if sql_parts:
        print(f"   🐍 Detected {len(sql_parts)} Python variable assignment(s) with SQL, extracting...")
        # Combine all SQL queries with a semicolon separator
        content = '\n;\n'.join(sql_parts)
    
    # Try to parse as SQL first
    sql_objects = parse_sql_for_objects(content)
    
    if sql_objects:
        # File contains SQL - use existing SQL parsing logic (handles temp table exclusion)
        print("   📄 Detected SQL content in file, parsing with SQL parser...")
        return get_distinct_objects_from_sql(content, default_database, default_schema)
    
    # If not SQL, parse line by line for object names
    print("   📝 Parsing as line-delimited object names...")
    
    database = default_database or ''
    schema = default_schema or ''
    object_names = []
    
    lines = content.strip().split('\n')
    
    for line in lines:
        line = line.strip()
        
        # Skip empty lines and comments
        if not line or line.startswith('--') or line.startswith('#'):
            continue
        
        # Parse object reference (database.schema.object or schema.object or object)
        parts = line.split('.')
        
        if len(parts) == 3:
            # Fully qualified: database.schema.object
            if not database:
                database = parts[0]
            if not schema:
                schema = parts[1]
            object_names.append(parts[2])
        elif len(parts) == 2:
            # Schema qualified: schema.object
            if not schema:
                schema = parts[0]
            object_names.append(parts[1])
        elif len(parts) == 1:
            # Simple object name
            object_names.append(parts[0])
    
    # Remove duplicates while preserving order
    object_names = list(dict.fromkeys(object_names))
    
    # Apply database name mapping (PROD_UHCCD -> ZJX_PRD_UHCCD_DB)
    original_database = database
    database = map_database_name(database if database else default_database)
    
    # Use default schema if not found
    if not schema:
        schema = default_schema if default_schema else ''
    
    print(f"\n📋 Parsed .txt file and found:")
    if original_database and original_database != database:
        print(f"   Database (original): {original_database} -> Mapped to: {database}")
    else:
        print(f"   Database: {database if database else '(not specified, using default)'}")
    print(f"   Schema: {schema if schema else '(not specified, using default)'}")
    print(f"   Objects ({len(object_names)}): {', '.join(object_names)}")
    
    # Log to file as well
    if object_names:
        print(f"\n📝 Extracted {len(object_names)} object(s) from .txt file:")
        for idx, obj in enumerate(object_names, 1):
            print(f"   {idx}. {obj}")
    
    return database, schema, object_names


def get_distinct_objects_from_sql(sql_code: str, default_database: str = None, default_schema: str = None) -> Tuple[str, str, List[str]]:
    """
    Parse SQL code and return distinct database, schema, and list of object names.
    
    This function extracts all object references from SQL code and consolidates them
    into a single database, schema, and list of distinct object names.
    
    IMPORTANT: Temporary objects are automatically excluded from the results and logged separately:
    - Temporary tables (CREATE TEMPORARY TABLE)
    - CTEs/Common Table Expressions (WITH cte_name AS)
    - Subquery aliases (SELECT ... FROM (...) AS alias)
    
    Logic:
    ======
    1. Parse SQL to find all object references
    2. Extract temporary tables, CTEs, and subquery aliases to exclude them
    3. Use the first found database (or default if not found)
    4. Use the first found schema (or default if not found)
    5. Collect all distinct object names (excluding temporary objects)
    6. Return (database, schema, object_list)
    
    Args:
        sql_code (str): SQL code to parse
        default_database (str): Default database if not found in SQL
        default_schema (str): Default schema if not found in SQL
    
    Returns:
        Tuple[str, str, List[str]]: (database, schema, list of object names)
    
    Example:
        >>> sql = '''
        ... CREATE TEMPORARY TABLE temp_data AS
        ... SELECT * FROM ZJX_PRD_UHCCD_DB.UDWBASESECUREVIEW1.TABLE1
        ... JOIN UDWBASESECUREVIEW1.TABLE2 ON ...
        ... '''
        >>> db, schema, objects = get_distinct_objects_from_sql(sql)
        >>> print(f"Objects: {objects}")  # temp_data excluded
        Objects: ['TABLE1', 'TABLE2']
    """
    # First, extract temporary table names, CTEs, and subquery aliases to exclude them
    temp_tables = extract_temp_table_names(sql_code)
    cte_names = extract_cte_names(sql_code)
    subquery_aliases = extract_subquery_aliases(sql_code)
    
    # Combine all exclusions
    exclusions_upper = [t.upper() for t in (temp_tables + cte_names + subquery_aliases)]
    
    # Log exclusions by category
    if temp_tables:
        print(f"\n🚫 Temporary Tables Detected (will be EXCLUDED from lineage):")
        for idx, temp_table in enumerate(temp_tables, 1):
            print(f"   {idx}. {temp_table}")
        print(f"   ℹ️  {len(temp_tables)} temporary table(s) will be excluded from processing")
    
    if cte_names:
        print(f"\n🚫 CTEs (Common Table Expressions) Detected (will be EXCLUDED from lineage):")
        for idx, cte in enumerate(cte_names, 1):
            print(f"   {idx}. {cte}")
        print(f"   ℹ️  {len(cte_names)} CTE(s) will be excluded from processing")
    
    if subquery_aliases:
        print(f"\n🚫 Subquery Aliases Detected (will be EXCLUDED from lineage):")
        for idx, alias in enumerate(subquery_aliases, 1):
            print(f"   {idx}. {alias}")
        print(f"   ℹ️  {len(subquery_aliases)} subquery alias(es) will be excluded from processing")
    
    # Parse SQL to find all object references
    objects = parse_sql_for_objects(sql_code)
    
    if not objects:
        print("⚠️  No objects found in SQL code")
        return default_database or '', default_schema or '', []
    
    # Extract database (use first non-empty database found, or default)
    database = default_database or ''
    for obj in objects:
        if obj['database']:
            database = obj['database']
            break
    
    # Extract schema (use first non-empty schema found, or default)
    schema = default_schema or ''
    for obj in objects:
        if obj['schema']:
            schema = obj['schema']
            break
    
    # Extract distinct object names and filter out temporary tables, CTEs, and aliases
    all_object_names = [obj['object'] for obj in objects]
    object_names = []
    excluded_objects = []
    
    for obj_name in all_object_names:
        if obj_name.upper() in exclusions_upper:
            excluded_objects.append(obj_name)
        else:
            if obj_name not in object_names:  # Preserve order, remove duplicates
                object_names.append(obj_name)
    
    if excluded_objects:
        excluded_count = len(set(excluded_objects))
        total_exclusions = len(temp_tables) + len(cte_names) + len(subquery_aliases)
        print(f"\n⚠️  Excluded {excluded_count} reference(s) to temporary objects (temp tables/CTEs/aliases) from objects list")
    
    # Apply database name mapping (PROD_UHCCD -> ZJX_PRD_UHCCD_DB)
    original_database = database
    database = map_database_name(database if database else default_database)
    
    # Use default schema if not found
    if not schema:
        schema = default_schema if default_schema else ''
    
    print(f"\n📋 Parsed SQL and found:")
    if original_database and original_database != database:
        print(f"   Database (original): {original_database} -> Mapped to: {database}")
    else:
        print(f"   Database: {database if database else '(not specified, using default)'}")
    print(f"   Schema: {schema if schema else '(not specified, using default)'}")
    print(f"   Objects ({len(object_names)}): {', '.join(object_names)}")
    
    # Log individual objects (excluding temp tables, CTEs, and aliases)
    if object_names:
        print(f"\n📝 Extracted {len(object_names)} REAL object(s) from SQL (excluding temp tables/CTEs/aliases):")
        for idx, obj in enumerate(object_names, 1):
            print(f"   {idx}. {obj}")
    else:
        print(f"\n⚠️  No real database objects found (all references were temporary objects)")
    
    return database, schema, object_names


def _parse_sys_args():
    """Parse CLI args using sys.argv.
    
    USAGE MODES:
    ============
    
    Mode 1: Single Object (Original behavior)
      python script.py <database> <schema> <object>
    Example:
      python script.py ZJX_PRD_UHCCD_DB UDWBASESECUREVIEW1 ADJD_MCE_MAIN_F
    
    Mode 2: Multiple Objects (Batch Processing)
      python script.py <database> <schema> <object1> <object2> <object3> ...
    Example:
      python script.py ZJX_PRD_UHCCD_DB UDWBASESECUREVIEW1 ADJD_MCE_MAIN_F ADJD_MCE_PROV USER_SECUR_KEY_SK
    
    Mode 3: SQL File Input (Parse SQL to extract objects)
      python script.py --sql <sql_file_path> [--database <db>] [--schema <schema>]
    Example:
      python script.py --sql query.sql --database ZJX_PRD_UHCCD_DB --schema UDWBASESECUREVIEW1
      python script.py --sql query.sql  (uses database/schema from SQL if available)
    
    Mode 4: TXT File Input - Lineage Only
      python script.py --txt <txt_file_path> --mode lineage-only
    Example:
      python script.py --txt source/views.txt --mode lineage-only
      (Runs lineage analysis and cluster key check only, no query execution)
    
    Mode 5: TXT File Input - Lineage with Query Execution
      python script.py --txt <txt_file_path> --mode lineage-with-execution --platform <snowflake|databricks> --timeout <minutes>
    Example:
      python script.py --txt source/views.txt --mode lineage-with-execution --platform snowflake --timeout 5
      (Runs lineage, cluster key check, and executes query with timeout)
    
    Mode 6: Auto-detect TXT file in source folder
      python script.py --auto-source [--database <db>] [--schema <schema>]
    Example:
      python script.py --auto-source  (automatically finds and processes .txt files in source/)
    
    Mode 7: No Arguments (Use Defaults)
      python script.py
    
    Returns:
      Tuple of (database, schema, object_list, txt_file_path, timeout_minutes, platform, execution_mode)
      - database: str - Snowflake database name
      - schema: str - Snowflake schema name
      - object_list: List[str] - List of one or more object names to process
      - txt_file_path: str - Path to txt file (if --txt mode)
      - timeout_minutes: int - Timeout in minutes (if execution mode)
      - platform: str - Platform for query execution (snowflake/databricks)
      - execution_mode: str - 'lineage-only' or 'lineage-with-execution'
    """
    # Check for auto-source mode (automatically find .txt files in source folder)
    if len(sys.argv) >= 2 and sys.argv[1] == '--auto-source':
        source_folder = os.path.join(os.path.dirname(__file__), 'source')
        
        if not os.path.exists(source_folder):
            print(f"❌ Error: Source folder not found: {source_folder}")
            sys.exit(-1)
        
        # Find all .txt files in source folder
        txt_files = [f for f in os.listdir(source_folder) if f.endswith('.txt')]
        
        if not txt_files:
            print(f"❌ Error: No .txt files found in {source_folder}")
            sys.exit(-1)
        
        if len(txt_files) > 1:
            print(f"⚠️  Multiple .txt files found in source folder:")
            for i, f in enumerate(txt_files, 1):
                print(f"   {i}. {f}")
            print(f"\n   Using first file: {txt_files[0]}")
        
        txt_file_path = os.path.join(source_folder, txt_files[0])
        
        # Parse optional database and schema flags
        default_db = None
        default_schema = None
        
        i = 2
        while i < len(sys.argv):
            if sys.argv[i] == '--database':
                if i + 1 < len(sys.argv):
                    default_db = sys.argv[i + 1]
                    i += 2
                else:
                    print("❌ Error: --database flag requires a value")
                    sys.exit(-1)
            elif sys.argv[i] == '--schema':
                if i + 1 < len(sys.argv):
                    default_schema = sys.argv[i + 1]
                    i += 2
                else:
                    print("❌ Error: --schema flag requires a value")
                    sys.exit(-1)
            else:
                print(f"⚠️  Unknown argument: {sys.argv[i]}")
                i += 1
        
        # Parse the .txt file
        db, schema, objects = get_objects_from_txt_file(txt_file_path, default_db, default_schema)
        
        # Validate that we have database and schema
        if not db:
            print("❌ Error: Database not found. Please specify with --database flag")
            sys.exit(-1)
        if not schema:
            print("❌ Error: Schema not found. Please specify with --schema flag")
            sys.exit(-1)
        if not objects:
            print("❌ Error: No objects found in file")
            sys.exit(-1)
        
        return db, schema, objects, None, None, None, None, None
    
    # Check for TXT file mode
    elif len(sys.argv) >= 2 and sys.argv[1] == '--txt':
        # TXT file mode: parse .txt file to extract objects
        if len(sys.argv) < 3:
            print("❌ Error: --txt flag requires a file path")
            print("Usage: python script.py --txt <txt_file_path> --mode <lineage-only|lineage-with-execution> [--platform <snowflake|databricks>] [--timeout <minutes>] [--check-platform <snowflake|databricks>]")
            sys.exit(-1)
        
        txt_file_path = sys.argv[2]
        
        # Parse optional flags
        execution_mode = None  # 'lineage-only' or 'lineage-with-execution'
        timeout_minutes = None
        platform = None
        check_platform = None  # Platform to check cluster keys against (snowflake or databricks)
        
        i = 3
        while i < len(sys.argv):
            if sys.argv[i] == '--mode':
                if i + 1 < len(sys.argv):
                    mode_value = sys.argv[i + 1]
                    if mode_value in ['lineage-only', 'lineage-with-execution']:
                        execution_mode = mode_value
                    else:
                        print("❌ Error: --mode must be 'lineage-only' or 'lineage-with-execution'")
                        sys.exit(-1)
                    i += 2
                else:
                    print("❌ Error: --mode flag requires a value")
                    sys.exit(-1)
            elif sys.argv[i] == '--platform':
                if i + 1 < len(sys.argv):
                    platform_value = sys.argv[i + 1].lower()
                    if platform_value in ['snowflake', 'databricks']:
                        platform = platform_value
                    else:
                        print("❌ Error: --platform must be 'snowflake' or 'databricks'")
                        sys.exit(-1)
                    i += 2
                else:
                    print("❌ Error: --platform flag requires a value")
                    sys.exit(-1)
            elif sys.argv[i] == '--check-platform':
                if i + 1 < len(sys.argv):
                    check_platform_value = sys.argv[i + 1].lower()
                    if check_platform_value in ['snowflake', 'databricks']:
                        check_platform = check_platform_value
                    else:
                        print("❌ Error: --check-platform must be 'snowflake' or 'databricks'")
                        sys.exit(-1)
                    i += 2
                else:
                    print("❌ Error: --check-platform flag requires a value")
                    sys.exit(-1)
            elif sys.argv[i] == '--timeout':
                if i + 1 < len(sys.argv):
                    try:
                        timeout_minutes = int(sys.argv[i + 1])
                    except ValueError:
                        print("❌ Error: --timeout value must be an integer")
                        sys.exit(-1)
                    i += 2
                else:
                    print("❌ Error: --timeout flag requires a value")
                    sys.exit(-1)
            else:
                print(f"⚠️  Unknown argument: {sys.argv[i]}")
                i += 1
        
        # If mode not provided, prompt user interactively
        if not execution_mode:
            print("\n" + "="*80)
            print("🔧 MODE SELECTION")
            print("="*80)
            print("Please select execution mode:")
            print("  1. lineage-only          - Lineage analysis + cluster key check only (no query execution)")
            print("  2. lineage-with-execution - Lineage + cluster key check + query execution")
            
            while True:
                mode_choice = input("\nEnter your choice (1 or 2): ").strip()
                if mode_choice == '1':
                    execution_mode = 'lineage-only'
                    break
                elif mode_choice == '2':
                    execution_mode = 'lineage-with-execution'
                    break
                else:
                    print("❌ Invalid choice. Please enter 1 or 2")
        
        # Validate requirements for lineage-with-execution mode
        if execution_mode == 'lineage-with-execution':
            # Prompt for platform if not provided
            if not platform:
                print("\n" + "="*80)
                print("☁️  PLATFORM SELECTION")
                print("="*80)
                print("Please select query execution platform:")
                print("  1. snowflake")
                print("  2. databricks")
                
                while True:
                    platform_choice = input("\nEnter your choice (1 or 2): ").strip()
                    if platform_choice == '1':
                        platform = 'snowflake'
                        break
                    elif platform_choice == '2':
                        platform = 'databricks'
                        break
                    else:
                        print("❌ Invalid choice. Please enter 1 or 2")
            
            # Prompt for timeout if not provided
            if not timeout_minutes:
                print("\n" + "="*80)
                print("⏱️  TIMEOUT CONFIGURATION")
                print("="*80)
                
                while True:
                    timeout_input = input("Enter query timeout in minutes (e.g., 5, 10): ").strip()
                    try:
                        timeout_minutes = int(timeout_input)
                        if timeout_minutes > 0:
                            break
                        else:
                            print("❌ Timeout must be greater than 0")
                    except ValueError:
                        print("❌ Please enter a valid number")
        
        # For lineage-only mode, timeout and platform are not needed (set to None if provided)
        if execution_mode == 'lineage-only':
            if timeout_minutes:
                print("ℹ️  Note: Timeout parameter ignored in lineage-only mode")
                timeout_minutes = None
            if platform:
                print("ℹ️  Note: Platform parameter ignored in lineage-only mode")
                platform = None

        
        # ALWAYS use Snowflake database for lineage analysis (regardless of platform selection)
        default_db = os.getenv("s_database")
        if not default_db:
            print("❌ Error: s_database not set in .env file")
            sys.exit(-1)
        
        print("\n" + "="*80)
        print("🚀 MODE CONFIGURATION")
        print("="*80)
        print(f"📄 File: {txt_file_path}")
        print(f"🔧 Mode: {execution_mode}")
        print(f"📊 Lineage Database: {default_db} (Snowflake)")
        
        if execution_mode == 'lineage-with-execution':
            print(f"☁️  Query Execution Platform: {platform.upper()}")
            print(f"⏱️  Timeout: {timeout_minutes} minutes")
        
        print("="*80 + "\n")
        
        # Parse the .txt file (schema will be extracted from file, or use default)
        default_schema = "UDWBASESECUREVIEW1"  # Default schema if not found in file
        db, schema, objects = get_objects_from_txt_file(txt_file_path, default_db, default_schema)
        
        # Use default schema if not found in file
        if not schema:
            schema = default_schema
            print(f"ℹ️  Schema not found in file, using default: {schema}")
        
        if not objects:
            print("❌ Error: No objects found in file")
            sys.exit(-1)
        
        # Return all values including execution_mode and check_platform
        return db, schema, objects, txt_file_path, timeout_minutes, platform, execution_mode, check_platform
    
    # Check for SQL file mode
    elif len(sys.argv) >= 2 and sys.argv[1] == '--sql':
        # SQL file mode: parse SQL file to extract objects OR execute query
        if len(sys.argv) < 3:
            print("❌ Error: --sql flag requires a file path")
            print("Usage: python script.py --sql <sql_file_path> [--database <db>] [--schema <schema>] [--timeout <minutes>] [--platform <snowflake|databricks>]")
            sys.exit(-1)
        
        sql_file_path = sys.argv[2]
        
        # Check if file exists
        if not os.path.exists(sql_file_path):
            print(f"❌ Error: SQL file not found: {sql_file_path}")
            sys.exit(-1)
        
        # Parse optional flags
        default_db = None
        default_schema = None
        timeout_minutes = None
        platform = None
        
        i = 3
        while i < len(sys.argv):
            if sys.argv[i] == '--database':
                if i + 1 < len(sys.argv):
                    default_db = sys.argv[i + 1]
                    i += 2
                else:
                    print("❌ Error: --database flag requires a value")
                    sys.exit(-1)
            elif sys.argv[i] == '--schema':
                if i + 1 < len(sys.argv):
                    default_schema = sys.argv[i + 1]
                    i += 2
                else:
                    print("❌ Error: --schema flag requires a value")
                    sys.exit(-1)
            elif sys.argv[i] == '--timeout':
                if i + 1 < len(sys.argv):
                    try:
                        timeout_minutes = int(sys.argv[i + 1])
                        i += 2
                    except ValueError:
                        print("❌ Error: --timeout value must be an integer (minutes)")
                        sys.exit(-1)
                else:
                    print("❌ Error: --timeout flag requires a value (minutes)")
                    sys.exit(-1)
            elif sys.argv[i] == '--platform':
                if i + 1 < len(sys.argv):
                    platform = sys.argv[i + 1].lower()
                    if platform not in ['snowflake', 'databricks']:
                        print("❌ Error: --platform must be 'snowflake' or 'databricks'")
                        sys.exit(-1)
                    i += 2
                else:
                    print("❌ Error: --platform flag requires a value (snowflake or databricks)")
                    sys.exit(-1)
            else:
                print(f"⚠️  Unknown argument: {sys.argv[i]}")
                i += 1
        
        # Ask for platform if timeout is present but platform not specified
        if timeout_minutes is not None and platform is None:
            print("\n☁️  Platform Selection:")
            print("   1. Snowflake")
            print("   2. Databricks")
            while True:
                choice = input("\nSelect platform (1 or 2): ").strip()
                if choice == '1':
                    platform = 'snowflake'
                    break
                elif choice == '2':
                    platform = 'databricks'
                    break
                else:
                    print("❌ Invalid choice. Please enter 1 or 2.")
            print(f"   Selected: {platform.upper()}")
        
        # Read SQL file for lineage analysis
        print(f"\n📂 Reading SQL file: {sql_file_path}")
        try:
            with open(sql_file_path, 'r', encoding='utf-8') as f:
                sql_code = f.read()
            print(f"   ✅ File loaded ({len(sql_code)} characters)")
        except Exception as e:
            print(f"❌ Error reading SQL file: {str(e)}")
            sys.exit(-1)
        
        # Parse SQL to extract objects
        db, schema, objects = get_distinct_objects_from_sql(sql_code, default_db, default_schema)
        
        # Validate that we have database and schema
        if not db:
            print("❌ Error: Database not found in SQL. Please specify with --database flag")
            sys.exit(-1)
        if not schema:
            print("❌ Error: Schema not found in SQL. Please specify with --schema flag")
            sys.exit(-1)
        if not objects:
            print("❌ Error: No objects found in SQL file")
            sys.exit(-1)
        
        return db, schema, objects, None, None, None, None, None
    
    elif len(sys.argv) == 1:
        # No args: use defaults
        default_db = "ZJX_PRD_UHCCD_DB"
        default_schema = "UDWBASESECUREVIEW1"
        default_objects = ["ADJD_MCE_MAIN_F"]
        print(f"No CLI args supplied. Using defaults:")
        print(f"  Database: {default_db}")
        print(f"  Schema: {default_schema}")
        print(f"  Objects: {', '.join(default_objects)}")
        return default_db, default_schema, default_objects, None, None, None, None, None
    
    elif len(sys.argv) >= 4:
        # Multiple objects: database schema object1 [object2 object3 ...]
        db = sys.argv[1]
        schema = sys.argv[2]
        objects = sys.argv[3:]  # All remaining arguments are object names
        
        print(f"\nCLI Arguments parsed:")
        print(f"  Database: {db}")
        print(f"  Schema: {schema}")
        print(f"  Objects ({len(objects)}): {', '.join(objects)}")
        
        return db, schema, objects, None, None, None, None, None
    
    else:
        print("Usage: python script.py <database> <schema> <object1> [object2 object3 ...]")
        print("\nExamples:")
        print("  Single object:")
        print("    python script.py ZJX_PRD_UHCCD_DB UDWBASESECUREVIEW1 ADJD_MCE_MAIN_F")
        print("\n  Multiple objects:")
        print("    python script.py ZJX_PRD_UHCCD_DB UDWBASESECUREVIEW1 ADJD_MCE_MAIN_F ADJD_MCE_PROV USER_SECUR_KEY_SK")
        sys.exit(-1)


def process_objects_sequentially(database: str, schema: str, object_names: List[str]):
    """
    Process multiple object names sequentially.
    
    This method iterates through each object name and calls the main() function
    to perform lineage discovery, record counting, and Excel file generation for each object.
    
    Processing Flow:
    ================
    1. Validates input parameters
    2. Logs the batch processing session
    3. Iterates through each object name sequentially
    4. For each object:
       - Calls main() to process that object
       - Logs completion status
       - Handles exceptions gracefully
       - Continues to next object even if one fails
    5. Generates a summary report at the end
    
    Args:
        database (str): Snowflake database name (same for all objects)
        schema (str): Snowflake schema name (same for all objects)
        object_names (List[str]): List of object names to process
    
    Returns:
        Dict with processing results:
        {
            'total_objects': int,
            'successful': int,
            'failed': int,
            'results': List[Dict] - detailed results for each object
        }
    
    Example:
        >>> results = process_objects_sequentially(
        ...     database="ZJX_PRD_UHCCD_DB",
        ...     schema="UDWBASESECUREVIEW1",
        ...     object_names=["ADJD_MCE_MAIN_F", "ADJD_MCE_PROV", "USER_SECUR_KEY_SK"]
        ... )
        >>> print(f"Processed {results['successful']}/{results['total_objects']} objects successfully")
    """
    from datetime import datetime
    
    # Input validation
    if not database or not schema or not object_names:
        print("❌ Error: database, schema, and object_names are required")
        return {
            'total_objects': 0,
            'successful': 0,
            'failed': 0,
            'results': []
        }
    
    if not isinstance(object_names, (list, tuple)):
        print("❌ Error: object_names must be a list")
        return {
            'total_objects': 0,
            'successful': 0,
            'failed': 0,
            'results': []
        }
    
    total_objects = len(object_names)
    successful = 0
    failed = 0
    skipped = 0  # Objects with no dependencies
    results = []
    
    # Session start
    session_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    print("\n" + "="*80)
    print("🚀 BATCH PROCESSING SESSION STARTED")
    print("="*80)
    print(f"📅 Session ID: {session_timestamp}")
    print(f"🗄️  Database: {database}")
    print(f"📂 Schema: {schema}")
    print(f"📋 Objects to process: {total_objects}")
    print(f"   - {', '.join(object_names)}")
    print("="*80 + "\n")
    
    # Create a single Snowflake connection for all objects
    print("🔌 Establishing single Snowflake connection for all objects...")
    conn = get_snowflake_connection()
    cursor = conn.cursor()
    print("✅ Connected to Snowflake (will be reused for all objects)\n")
    
    # Process each object sequentially using the same connection
    for index, obj_name in enumerate(object_names, 1):
        obj_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        print("\n" + "-"*80)
        print(f"[{index}/{total_objects}] 🔄 Processing: {obj_name} (Started at {obj_timestamp})")
        print("-"*80 + "\n")
        
        try:
            # Call main() for this object with shared connection
            main(TARGET_DATABASE=database, TARGET_SCHEMA=schema, TARGET_OBJECT=obj_name, conn=conn, cursor=cursor)
            
            # Mark as successful
            successful += 1
            result_status = "✅ SUCCESS"
            result_entry = {
                'object_name': obj_name, 
                'status': 'SUCCESS',
                'timestamp': obj_timestamp,
                'error': None
            }
            
            print(f"\n{result_status}: {obj_name} processed successfully")
            
        except Exception as e:
            error_msg = str(e)
            
            # Check if this is a "no dependencies" case
            if "NO_DEPENDENCIES" in error_msg:
                skipped += 1
                result_status = "⏭️  SKIPPED"
                result_entry = {
                    'object_name': obj_name,
                    'status': 'SKIPPED',
                    'timestamp': obj_timestamp,
                    'error': 'No dependencies found (base table or no upstream lineage)'
                }
                print(f"\n{result_status}: {obj_name} - No dependencies found")
            else:
                # Mark as failed
                failed += 1
                result_status = "❌ FAILED"
                result_entry = {
                    'object_name': obj_name,
                    'status': 'FAILED',
                    'timestamp': obj_timestamp,
                    'error': error_msg
                }
                
                print(f"\n{result_status}: {obj_name}")
                print(f"   Error: {error_msg}")
                import traceback
                print(f"   Traceback: {traceback.format_exc()}")
        
        results.append(result_entry)
    
    # Close the shared connection after all objects are processed
    cursor.close()
    conn.close()
    print("\n🔌 Snowflake connection closed\n")
    
    # Session end - Summary report
    print("\n" + "="*80)
    print("📊 BATCH PROCESSING COMPLETED")
    print("="*80)
    print(f"📅 Session ID: {session_timestamp}")
    print(f"✅ Successful: {successful}/{total_objects}")
    print(f"⏭️  Skipped (No Dependencies): {skipped}/{total_objects}")
    print(f"❌ Failed: {failed}/{total_objects}")
    
    if successful == total_objects:
        print("\n🎉 All objects processed successfully!")
    elif successful == 0 and skipped == 0:
        print("\n⚠️  All objects failed to process. Please check the logs above.")
    elif successful > 0 or skipped > 0:
        print(f"\n📋 Summary: {successful} succeeded, {skipped} skipped (no dependencies), {failed} failed")
    
    # Detailed results
    if results:
        print("\n📝 Detailed Results:")
        successful_list = [r for r in results if r['status'] == 'SUCCESS']
        skipped_list = [r for r in results if r['status'] == 'SKIPPED']
        failed_list = [r for r in results if r['status'] == 'FAILED']
        
        if successful_list:
            print("\n✅ Successful Objects:")
            for i, result in enumerate(successful_list, 1):
                print(f"   {i}. {result['object_name']}")
        
        if skipped_list:
            print("\n⏭️  Skipped Objects (No Dependencies):")
            for i, result in enumerate(skipped_list, 1):
                print(f"   {i}. {result['object_name']} - {result['error']}")
        
        if failed_list:
            print("\n❌ Failed Objects:")
            for i, result in enumerate(failed_list, 1):
                print(f"   {i}. {result['object_name']}")
                print(f"      Error: {result['error'][:100]}...")
    
    print("="*80 + "\n")
    
    return {
        'total_objects': total_objects,
        'successful': successful,
        'skipped': skipped,
        'failed': failed,
        'results': results,
        'session_id': session_timestamp
    }


if __name__ == "__main__":
    # Parse arguments (handles --txt, --sql, or direct args)
    db, schema, object_list, txt_file_path, timeout_minutes, platform, execution_mode, check_platform = _parse_sys_args()
    
    # If check_platform not provided, prompt user
    if txt_file_path and not check_platform:
        print("\n" + "="*80)
        print("🔍 CLUSTER KEY CHECK PLATFORM SELECTION")
        print("="*80)
        print("Which platform's cluster/partition keys do you want to check?")
        print("  1. Snowflake  - Check SF_Clustering_Keys from TD_vs_SF_Comparison sheet")
        print("  2. Databricks - Check DB_clusteringColumns from Databricks_Info sheet")
        
        while True:
            check_choice = input("\nEnter your choice (1 or 2): ").strip()
            if check_choice == '1':
                check_platform = 'snowflake'
                break
            elif check_choice == '2':
                check_platform = 'databricks'
                break
            else:
                print("❌ Invalid choice. Please enter 1 or 2")
        
        print(f"   Selected: {check_platform.upper()}")
        print("="*80 + "\n")
    
    # If --txt mode with lineage-with-execution, run lineage analysis first then execute query
    if txt_file_path and execution_mode == 'lineage-with-execution':
        # Platform already selected in _parse_sys_args, no need to prompt again
        
        # Run lineage analysis (Snowflake) + query execution (selected platform)
        print("\n" + "="*80)
        print("🚀 MODE: LINEAGE ANALYSIS + CLUSTER KEY CHECK + QUERY EXECUTION")
        print("="*80)
        print(f"📄 File: {txt_file_path}")
        print(f"⏱️  Timeout: {timeout_minutes} minutes")
        print(f"☁️  Query Platform: {platform.upper()}")
        print(f"📊 Database: {db}")
        print(f"📊 Schema: {schema}")
        print(f"📊 Objects to analyze: {len(object_list)}")
        print("="*80 + "\n")
        
        # Step 1: Run lineage analysis (ALWAYS Snowflake)
        print("\n" + "="*80)
        print("STEP 1: RUNNING LINEAGE ANALYSIS (SNOWFLAKE)")
        print("="*80)
        print(f"📊 Database: {db}")
        print(f"📊 Schema: {schema}")
        print(f"📊 Objects to process: {len(object_list)}")
        print("\n📋 Object List:")
        for idx, obj in enumerate(object_list, 1):
            print(f"   {idx}. {obj}")
        print("="*80 + "\n")
        
        # Process objects sequentially (always uses Snowflake)
        batch_results = process_objects_sequentially(
            database=db,
            schema=schema,
            object_names=object_list
        )
        
        # Step 2: Execute query with timeout (uses selected platform)
        print("\n\n" + "="*80)
        print(f"STEP 2: EXECUTING QUERY IN {platform.upper()} WITH TIMEOUT")
        print("="*80)
        returned_query_id = process_query_file(txt_file_path, timeout_minutes, platform)
        
        # Display returned query ID
        if returned_query_id:
            print(f"\n🔑 Returned Query ID: {returned_query_id}")
            print(f"   You can use this Query ID in other scripts\n")
            
            # Step 2.5: Get execution plan (platform-specific)
            print("\n" + "="*80)
            print(f"STEP 2.5: DOWNLOADING EXECUTION PLAN ({platform.upper()})")
            print("="*80)
            
            # Extract query name from txt file path
            query_name = os.path.splitext(os.path.basename(txt_file_path))[0]
            print(f"📄 Reading query from: {txt_file_path}")
            print(f"📝 Query name: {query_name}")
            
            # Create Execution_plan folder if it doesn't exist
            execution_plan_dir = "Execution_plan"
            if not os.path.exists(execution_plan_dir):
                os.makedirs(execution_plan_dir)
                print(f"📁 Created directory: {execution_plan_dir}")
            else:
                print(f"📁 Using existing directory: {execution_plan_dir}")
            
            # Read the original query from file
            try:
                print(f"🔄 Loading SQL query from {txt_file_path}...")
                with open(txt_file_path, 'r') as f:
                    original_query = f.read().strip()
                print(f"✅ Query loaded successfully ({len(original_query)} characters)")
            except Exception as e:
                print(f"⚠️  Could not read query file for EXPLAIN: {e}")
                original_query = None
            
            if original_query:
                if platform == 'snowflake':
                    try:
                        print(f"🔌 Connecting to Snowflake...")
                        # Create new Snowflake connection for EXPLAIN
                        sf_conn = get_snowflake_connection()
                        cursor = sf_conn.cursor()
                        print(f"✅ Connected to Snowflake")
                        
                        # Execute EXPLAIN USING JSON
                        explain_query = f"EXPLAIN USING JSON {original_query}"
                        print(f"🔍 Executing: EXPLAIN USING JSON on query from {query_name}.txt...")
                        cursor.execute(explain_query)
                        
                        # Fetch the JSON result
                        explain_result = cursor.fetchone()
                        
                        if explain_result:
                            import json
                            # The result is typically in the first column
                            explain_json = explain_result[0]
                            
                            # Save to JSON file
                            json_filename = f"{execution_plan_dir}/{query_name}_execution_plan.json"
                            with open(json_filename, 'w') as json_file:
                                if isinstance(explain_json, str):
                                    json_file.write(explain_json)
                                else:
                                    json.dump(explain_json, json_file, indent=2)
                            
                            print(f"✅ Execution plan saved to: {json_filename}")
                        else:
                            print("⚠️  No execution plan returned")
                        
                        cursor.close()
                        sf_conn.close()
                        
                    except Exception as e:
                        print(f"❌ Error getting execution plan: {str(e)}")
                        import traceback
                        traceback.print_exc()
                
                elif platform == 'databricks':
                    try:
                        print(f"🔌 Connecting to Databricks...")
                        # Connect to Databricks
                        DATABRICKS_HOST = os.getenv("DATABRICKS_HOST")
                        DATABRICKS_HTTP_PATH = os.getenv("DATABRICKS_HTTP_PATH")
                        DATABRICKS_TOKEN = os.getenv("DATABRICKS_TOKEN")
                        
                        if not DATABRICKS_HOST or not DATABRICKS_HTTP_PATH or not DATABRICKS_TOKEN:
                            raise Exception("Databricks credentials not set in environment")
                        
                        import databricks.sql
                        db_conn = databricks.sql.connect(
                            server_hostname=DATABRICKS_HOST.replace("https://", ""),
                            http_path=DATABRICKS_HTTP_PATH,
                            access_token=DATABRICKS_TOKEN
                        )
                        print(f"✅ Connected to Databricks")
                        
                        db_cursor = db_conn.cursor()
                        
                        # Set catalog before executing EXPLAIN
                        print(f"🗄️  Setting catalog: USE CATALOG prod_uhccd...")
                        db_cursor.execute("USE CATALOG prod_uhccd;")
                        print(f"✅ Catalog set to prod_uhccd")
                        
                        # Execute EXPLAIN FORMATTED
                        explain_query = f"EXPLAIN FORMATTED {original_query}"
                        print(f"🔍 Executing: EXPLAIN FORMATTED on query from {query_name}.txt...")
                        db_cursor.execute(explain_query)
                        print(f"✅ EXPLAIN FORMATTED completed")
                        print(f"✅ EXPLAIN FORMATTED completed")
                        
                        # Fetch all results
                        print(f"📥 Fetching execution plan results...")
                        explain_results = db_cursor.fetchall()
                        column_names = [desc[0] for desc in db_cursor.description]
                        print(f"✅ Retrieved {len(explain_results)} rows")
                        
                        if explain_results:
                            import pandas as pd
                            df_explain = pd.DataFrame(explain_results, columns=column_names)
                            
                            # Save to CSV file
                            csv_filename = f"{execution_plan_dir}/{query_name}_execution_plan.csv"
                            print(f"💾 Saving execution plan to: {csv_filename}")
                            df_explain.to_csv(csv_filename, index=False)
                            
                            print(f"✅ Execution plan saved to: {csv_filename}")
                            print(f"📊 Plan rows: {len(df_explain)}")
                            print(f"📊 Plan columns: {', '.join(column_names)}")
                        else:
                            print("⚠️  No execution plan returned")
                        
                        print(f"🔌 Closing Databricks connection...")
                        db_cursor.close()
                        db_conn.close()
                        print(f"✅ Connection closed")
                        
                    except Exception as e:
                        print(f"❌ Error getting execution plan: {str(e)}")
                        import traceback
                        traceback.print_exc()
            
            print("="*80 + "\n")
            
            # Step 3: Get query operator stats (platform-specific)
            if platform == 'snowflake':
                print("\n" + "="*80)
                print("STEP 3: DOWNLOADING QUERY OPERATOR STATS (SNOWFLAKE)")
                print("="*80)
                print(f"📊 Query ID: {returned_query_id}")
                
                try:
                    # Create new Snowflake connection
                    sf_conn = get_snowflake_connection()
                    cursor = sf_conn.cursor()
                    
                    # Enhanced query with performance metrics
                    stats_query = f"""
WITH ranked AS (
    SELECT *,
           execution_time_breakdown:"overall_percentage"::FLOAT AS overall_pct
    FROM TABLE(GET_QUERY_OPERATOR_STATS('{returned_query_id}'))
)
SELECT
    operator_id,
    operator_type,
    parent_operators,
    operator_attributes,
    operator_statistics,
    /* Extract bytes from the JSON string, then convert to GiB and GB */
    ROUND(
        (TRY_PARSE_JSON(operator_statistics:"spilling"::STRING):"bytes_spilled_local_storage"::NUMBER)
        / POWER(1024, 3)
    , 2) AS spilled_gib,
    TRY_PARSE_JSON(operator_statistics:"pruning"::STRING):"partitions_scanned" :: NUMBER as partitions_scanned,
    TRY_PARSE_JSON(operator_statistics:"pruning"::STRING):"partitions_total" :: NUMBER as partitions_total,
    execution_time_breakdown
FROM ranked
ORDER BY overall_pct DESC
LIMIT 10
"""
                    print(f"Executing enhanced query profile analysis...")
                    cursor.execute(stats_query)
                    
                    # Fetch results
                    stats_results = cursor.fetchall()
                    column_names = [desc[0] for desc in cursor.description]
                    
                    if stats_results:
                        # Create DataFrame
                        import pandas as pd
                        df_stats = pd.DataFrame(stats_results, columns=column_names)
                        
                        # Extract query name from txt file path (e.g., source/Ashitosh.txt -> Ashitosh)
                        query_name = os.path.splitext(os.path.basename(txt_file_path))[0]
                        csv_filename = f"Output/{query_name}_queryprofile.csv"
                        
                        # Save to CSV
                        df_stats.to_csv(csv_filename, index=False)
                        print(f"✅ Query profile saved to: {csv_filename}")
                        print(f"📊 Top {len(df_stats)} operators by execution time")
                        
                        print("="*80 + "\n")
                        
                    else:
                        print("⚠️  No operator stats returned")
                    
                    cursor.close()
                    sf_conn.close()
                    
                except Exception as e:
                    print(f"❌ Error getting query operator stats: {str(e)}")
                    import traceback
                    traceback.print_exc()
                
                print("="*80 + "\n")
                
            elif platform == 'databricks':
                print("\n" + "="*80)
                print("STEP 3: DOWNLOADING QUERY PROFILE (DATABRICKS)")
                print("="*80)
                print(f"📊 Query ID: {returned_query_id}")
                
                try:
                    # Get Databricks connection
                    if not DATABRICKS_HOST or not DATABRICKS_HTTP_PATH or not DATABRICKS_TOKEN:
                        raise Exception("Databricks credentials not set in environment")
                    
                    import databricks.sql
                    db_conn = databricks.sql.connect(
                        server_hostname=DATABRICKS_HOST.replace("https://", ""),
                        http_path=DATABRICKS_HTTP_PATH,
                        access_token=DATABRICKS_TOKEN
                    )
                    
                    db_cursor = db_conn.cursor()
                    
                    # Query profile from system_catalog.query.query_history
                    profile_query = f"""
                    SELECT 
                        execution_status,
                        statement_type,
                        execution_duration_ms,
                        start_time,
                        end_time,
                        read_bytes,
                        read_files,
                        read_partitions,
                        read_rows,
                        produced_rows,
                        written_bytes,
                        written_files,
                        written_rows,
                        workspace_name,
                        shuffle_read_bytes / (1024 * 1024 * 1024) AS shuffle_read_gb,
                        spilled_local_bytes / (1024 * 1024 * 1024) AS spilled_local_bytes_gb
                    FROM system_catalog.query.query_history
                    WHERE statement_id = '{returned_query_id}'
                    """
                    
                    print(f"Executing query profile analysis...")
                    db_cursor.execute(profile_query)
                    
                    # Fetch results
                    profile_results = db_cursor.fetchall()
                    column_names = [desc[0] for desc in db_cursor.description]
                    
                    if profile_results:
                        # Create DataFrame
                        import pandas as pd
                        df_profile = pd.DataFrame(profile_results, columns=column_names)
                        
                        # SAFETY CHECK: Verify we got the correct query type
                        if 'statement_type' in df_profile.columns:
                            statement_type = df_profile['statement_type'].values[0]
                            if statement_type != 'SELECT':
                                print(f"⚠️  WARNING: Query profile is for {statement_type} instead of SELECT!")
                                print(f"   This indicates a Query ID mismatch.")
                                print(f"   Query ID used: {returned_query_id}")
                                print(f"   Expected: User's SELECT query")
                                print(f"   Got: {statement_type} query")
                                logger.warning(f"Query profile mismatch: got {statement_type} instead of SELECT for query_id {returned_query_id}")
                        
                        # Extract query name from txt file path (e.g., source/sandeep.txt -> sandeep)
                        query_name = os.path.splitext(os.path.basename(txt_file_path))[0]
                        csv_filename = f"Output/{query_name}_queryprofile.csv"
                        
                        # Save to CSV
                        df_profile.to_csv(csv_filename, index=False)
                        print(f"✅ Query profile saved to: {csv_filename}")
                        print(f"📊 Profile records: {len(df_profile)}")
                        
                        # Display statement type for verification
                        if 'statement_type' in df_profile.columns:
                            print(f"✅ Statement Type: {df_profile['statement_type'].values[0]}")
                        
                        print("="*80 + "\n")
                        
                    else:
                        print("⚠️  No query profile returned")
                    
                    db_cursor.close()
                    db_conn.close()
                    
                except Exception as e:
                    print(f"❌ Error getting query profile: {str(e)}")
                    import traceback
                    traceback.print_exc()
                
                print("="*80 + "\n")
            
            # Step 4: Check for cluster keys/partition keys in query
            print("\n" + "="*80)
            print(f"STEP 4: CLUSTER KEY / PARTITION KEY ANALYSIS ({check_platform.upper()})")
            print("="*80)
            
            try:
                # Read the query from txt file
                with open(txt_file_path, 'r', encoding='utf-8') as f:
                    query_content = f.read().upper()
                
                # Extract query name from txt file path for log file
                query_name = os.path.splitext(os.path.basename(txt_file_path))[0]
                
                # Create cluster_output directory if it doesn't exist
                cluster_output_dir = "cluster_output"
                os.makedirs(cluster_output_dir, exist_ok=True)
                cluster_log_filename = os.path.join(cluster_output_dir, f"cluster_key_{query_name}.txt")
                
                # Find ALL Excel files generated from lineage analysis in this run
                import glob
                import time
                excel_files = glob.glob('Output/*.xlsx')
                
                if not excel_files:
                    print("⚠️  No Excel files found in Output/ directory")
                else:
                    # Get Excel files modified in the last 5 minutes (current run)
                    current_time = time.time()
                    recent_files = [f for f in excel_files if (current_time - os.path.getmtime(f)) < 300]
                    
                    if not recent_files:
                        # Fallback to just the most recent file
                        recent_files = [max(excel_files, key=os.path.getmtime)]
                    
                    print(f"📊 Reading {len(recent_files)} lineage file(s) from current run")
                    print(f"🔍 Checking {check_platform.upper()} cluster/partition keys")
                    
                    # Read cluster keys based on selected platform
                    # Use dict to track which objects each key came from
                    import pandas as pd
                    from collections import defaultdict
                    cluster_keys_map = defaultdict(set)  # key -> set of object names
                    
                    for excel_file in recent_files:
                        try:
                            if check_platform == 'snowflake':
                                # Read from TD_vs_SF_Comparison sheet -> SF_Clustering_Keys column
                                df_comparison = pd.read_excel(excel_file, sheet_name='TD_vs_SF_Comparison')
                                
                                for idx, row in df_comparison.iterrows():
                                    object_name = row.get('Object_Name', '')
                                    clustering_keys = row.get('SF_Clustering_Keys', '')
                                    if clustering_keys and isinstance(clustering_keys, str) and clustering_keys not in ['NONE', 'EXCLUDED', ''] and pd.notna(clustering_keys):
                                        # Split comma-separated keys and track source object
                                        for k in clustering_keys.split(','):
                                            if k.strip():
                                                key = k.strip().upper()
                                                cluster_keys_map[key].add(object_name)
                            
                            elif check_platform == 'databricks':
                                # Read from Databricks_Info sheet -> DB_clusteringColumns column
                                df_databricks = pd.read_excel(excel_file, sheet_name='Databricks_Info')
                                
                                for idx, row in df_databricks.iterrows():
                                    object_name = row.get('SF_Object', '')
                                    clustering_cols = row.get('DB_clusteringColumns', '')
                                    if clustering_cols and isinstance(clustering_cols, str) and clustering_cols not in ['NONE', 'EXCLUDED', ''] and pd.notna(clustering_cols):
                                        # Parse JSON array format like ["ORIG_SRC_SYS_CD","SRC_SYS_CD","FTP_CD"]
                                        import json
                                        try:
                                            # Try parsing as JSON array first
                                            if clustering_cols.startswith('[') and clustering_cols.endswith(']'):
                                                keys_list = json.loads(clustering_cols)
                                                for k in keys_list:
                                                    if k and isinstance(k, str) and k.strip():
                                                        key = k.strip().upper()
                                                        cluster_keys_map[key].add(object_name)
                                            else:
                                                # Fallback to comma-separated string
                                                for k in clustering_cols.split(','):
                                                    if k.strip():
                                                        key = k.strip().upper()
                                                        cluster_keys_map[key].add(object_name)
                                        except json.JSONDecodeError:
                                            # If JSON parsing fails, fall back to comma-separated
                                            for k in clustering_cols.split(','):
                                                if k.strip():
                                                    key = k.strip().upper()
                                                    cluster_keys_map[key].add(object_name)
                        
                        except Exception as e:
                            print(f"⚠️  Error reading {excel_file}: {e}")
                    
                    if cluster_keys_map:
                        # Prepare output content
                        output_lines = []
                        output_lines.append(f"🔑 Found {len(cluster_keys_map)} unique cluster/partition keys in lineage:")
                        
                        # Analyze query structure to identify JOIN, WHERE, and SELECT sections
                        def analyze_key_usage(key, query_text):
                            """Detect if key is used in JOIN, WHERE, or SELECT clauses"""
                            import re
                            
                            # Convert to uppercase for case-insensitive matching
                            query_upper = query_text.upper()
                            key_pattern = r'\b' + re.escape(key) + r'\b'
                            
                            # Find all matches of the key
                            matches = list(re.finditer(key_pattern, query_upper))
                            
                            in_join = False
                            in_where = False
                            in_select = False
                            
                            for match in matches:
                                pos = match.start()
                                
                                # Get context before the key (up to 500 chars)
                                context_start = max(0, pos - 500)
                                context = query_upper[context_start:pos]
                                
                                # Check if in JOIN clause
                                # Look for JOIN keywords before the key
                                if re.search(r'\b(INNER\s+JOIN|LEFT\s+JOIN|RIGHT\s+JOIN|FULL\s+JOIN|JOIN|LEFT\s+OUTER\s+JOIN|RIGHT\s+OUTER\s+JOIN)\b(?!.*\bWHERE\b)(?!.*\bSELECT\b)', context):
                                    # Check if it's in ON clause (after JOIN and before WHERE/GROUP BY/ORDER BY)
                                    extended_context = query_upper[context_start:min(len(query_upper), pos + 200)]
                                    if re.search(r'\bON\b', extended_context):
                                        in_join = True
                                
                                # Check if in WHERE clause
                                # Look for WHERE keyword before the key and ensure no SELECT after WHERE
                                if re.search(r'\bWHERE\b(?!.*\bSELECT\b)', context):
                                    # Make sure we're not in a subquery's SELECT
                                    # Check if there's a GROUP BY, ORDER BY, or HAVING after WHERE but before our key
                                    remaining_context = query_upper[pos:min(len(query_upper), pos + 100)]
                                    # If we see GROUP BY/ORDER BY/HAVING right after, we might be past WHERE clause
                                    if not re.search(r'^\s*(GROUP\s+BY|ORDER\s+BY|HAVING|LIMIT|\))', remaining_context):
                                        in_where = True
                                
                                # Check if in SELECT clause
                                # Look for SELECT keyword before the key
                                if re.search(r'\bSELECT\b(?!.*\bFROM\b)', context):
                                    # We're in SELECT if SELECT is before key and no FROM between them
                                    in_select = True
                                elif re.search(r'\bSELECT\b', context):
                                    # Check if we're between SELECT and FROM (select list)
                                    from_match = re.search(r'\bFROM\b', context)
                                    select_match = re.search(r'\bSELECT\b', context)
                                    if select_match and from_match:
                                        # If FROM comes after SELECT, we might be in SELECT list
                                        if select_match.start() > from_match.start():
                                            # SELECT is more recent than FROM, likely in a subquery SELECT
                                            remaining = query_upper[pos:min(len(query_upper), pos + 50)]
                                            if not re.search(r'^\s*FROM\b', remaining):
                                                in_select = True
                            
                            return in_join, in_where, in_select
                        
                        # Check which keys are present in the query using word boundary matching
                        keys_in_query = []
                        keys_not_in_query = []
                        
                        import re
                        for key in sorted(cluster_keys_map.keys()):
                            # Use word boundary regex to match whole words only
                            pattern = r'\b' + re.escape(key) + r'\b'
                            if re.search(pattern, query_content):
                                keys_in_query.append(key)
                            else:
                                keys_not_in_query.append(key)
                        
                        # Build output for keys in query
                        if keys_in_query:
                            output_lines.append(f"\n✅ Cluster/Partition Keys PRESENT in Query ({len(keys_in_query)}):")
                            for key in keys_in_query:
                                sources = ', '.join(sorted(cluster_keys_map[key]))
                                in_join, in_where, in_select = analyze_key_usage(key, query_content)
                                usage = f"(Join: {'Y' if in_join else 'N'}, Where: {'Y' if in_where else 'N'}, Select: {'Y' if in_select else 'N'})"
                                output_lines.append(f"   {key} (from: {sources}) {usage}")
                        
                        # Build output for keys not in query
                        if keys_not_in_query:
                            output_lines.append(f"\n⚠️  Cluster/Partition Keys NOT FOUND in Query ({len(keys_not_in_query)}):")
                            for key in keys_not_in_query:
                                sources = ', '.join(sorted(cluster_keys_map[key]))
                                output_lines.append(f"   {key} (from: {sources})")
                        
                        # Summary
                        output_lines.append(f"\n📈 Summary:")
                        output_lines.append(f"   Total cluster keys identified: {len(cluster_keys_map)}")
                        output_lines.append(f"   Keys used in query: {len(keys_in_query)} ({len(keys_in_query)*100//len(cluster_keys_map)}%)")
                        output_lines.append(f"   Keys missing from query: {len(keys_not_in_query)} ({len(keys_not_in_query)*100//len(cluster_keys_map)}%)")
                        
                        if len(keys_not_in_query) > 0:
                            output_lines.append(f"\n💡 Recommendation: Consider adding the missing cluster/partition keys to your WHERE clause for better performance")
                        
                        # Write to log file
                        with open(cluster_log_filename, 'w', encoding='utf-8') as log_file:
                            log_file.write('\n'.join(output_lines))
                        
                        # Print to console
                        print('\n'.join(output_lines))
                        print(f"\n📝 Cluster key analysis saved to: {cluster_log_filename}")
                    else:
                        output_msg = "ℹ️  No cluster keys or partition keys found in the lineage analysis"
                        print(output_msg)
                        
                        # Write to log file even if no keys found
                        with open(cluster_log_filename, 'w', encoding='utf-8') as log_file:
                            log_file.write(output_msg)
                        print(f"\n📝 Cluster key analysis saved to: {cluster_log_filename}")
                    
            except Exception as e:
                print(f"❌ Error analyzing cluster keys: {str(e)}")
                import traceback
                traceback.print_exc()
            
            print("="*80 + "\n")
    
    # If --txt mode with lineage-only, run lineage analysis and cluster key check without query execution
    elif txt_file_path and execution_mode == 'lineage-only':
        print("\n" + "="*80)
        print("🚀 MODE: LINEAGE ANALYSIS + CLUSTER KEY CHECK ONLY")
        print("="*80)
        print(f"📄 File: {txt_file_path}")
        print(f"📊 Database: {db}")
        print(f"📊 Schema: {schema}")
        print(f"📊 Objects to process: {len(object_list)}")
        print("\n📋 Object List:")
        for idx, obj in enumerate(object_list, 1):
            print(f"   {idx}. {obj}")
        print("="*80 + "\n")
        
        # Step 1: Run lineage analysis (ALWAYS Snowflake)
        print("\n" + "="*80)
        print("STEP 1: RUNNING LINEAGE ANALYSIS (SNOWFLAKE)")
        print("="*80)
        print(f"📊 Database: {db}")
        print(f"📊 Schema: {schema}")
        print(f"📊 Objects to process: {len(object_list)}")
        print("="*80 + "\n")
        
        # Process objects sequentially (always uses Snowflake)
        batch_results = process_objects_sequentially(
            database=db,
            schema=schema,
            object_names=object_list
        )
        
        # Step 2: Check for cluster keys/partition keys in query
        print("\n" + "="*80)
        print(f"STEP 2: CLUSTER KEY / PARTITION KEY ANALYSIS ({check_platform.upper()})")
        print("="*80)
        
        try:
            # Read the query from txt file
            with open(txt_file_path, 'r', encoding='utf-8') as f:
                query_content = f.read().upper()
            
            # Extract query name from txt file path for log file
            query_name = os.path.splitext(os.path.basename(txt_file_path))[0]
            
            # Create cluster_output directory if it doesn't exist
            cluster_output_dir = "cluster_output"
            os.makedirs(cluster_output_dir, exist_ok=True)
            cluster_log_filename = os.path.join(cluster_output_dir, f"cluster_key_{query_name}.txt")
            
            # Find ALL Excel files generated from lineage analysis in this run
            import glob
            import time
            excel_files = glob.glob('Output/*.xlsx')
            
            if not excel_files:
                print("⚠️  No Excel files found in Output/ directory")
            else:
                # Get Excel files modified in the last 20 minutes (current run)
                # Increased from 5 minutes to 20 minutes to capture all files from multi-object lineage runs
                current_time = time.time()
                recent_files = [f for f in excel_files if (current_time - os.path.getmtime(f)) < 1200]
                
                if not recent_files:
                    # Fallback to just the most recent file
                    recent_files = [max(excel_files, key=os.path.getmtime)]
                
                print(f"📊 Reading {len(recent_files)} lineage file(s) from current run")
                print(f"🔍 Checking {check_platform.upper()} cluster/partition keys")
                print(f"\n📂 Excel files being analyzed:")
                for idx, f in enumerate(recent_files, 1):
                    import os
                    basename = os.path.basename(f)
                    print(f"   {idx}. {basename}")
                
                # Read cluster keys based on selected platform
                # Use dict to track which objects each key came from
                import pandas as pd
                from collections import defaultdict
                cluster_keys_map = defaultdict(set)  # key -> set of object names
                
                for excel_file in recent_files:
                    try:
                        if check_platform == 'snowflake':
                            # Read from TD_vs_SF_Comparison sheet -> SF_Clustering_Keys column
                            df_comparison = pd.read_excel(excel_file, sheet_name='TD_vs_SF_Comparison')
                            
                            for idx, row in df_comparison.iterrows():
                                object_name = row.get('Object_Name', '')
                                clustering_keys = row.get('SF_Clustering_Keys', '')
                                if clustering_keys and isinstance(clustering_keys, str) and clustering_keys not in ['NONE', 'EXCLUDED', ''] and pd.notna(clustering_keys):
                                    # Split comma-separated keys and track source object
                                    for k in clustering_keys.split(','):
                                        if k.strip():
                                            key = k.strip().upper()
                                            cluster_keys_map[key].add(object_name)
                        
                        elif check_platform == 'databricks':
                            # Read from Databricks_Info sheet -> DB_clusteringColumns column
                            df_databricks = pd.read_excel(excel_file, sheet_name='Databricks_Info')
                            
                            for idx, row in df_databricks.iterrows():
                                object_name = row.get('SF_Object', '')
                                clustering_cols = row.get('DB_clusteringColumns', '')
                                if clustering_cols and isinstance(clustering_cols, str) and clustering_cols not in ['NONE', 'EXCLUDED', ''] and pd.notna(clustering_cols):
                                    # Parse JSON array format like ["ORIG_SRC_SYS_CD","SRC_SYS_CD","FTP_CD"]
                                    import json
                                    try:
                                        # Try parsing as JSON array first
                                        if clustering_cols.startswith('[') and clustering_cols.endswith(']'):
                                            keys_list = json.loads(clustering_cols)
                                            for k in keys_list:
                                                if k and isinstance(k, str) and k.strip():
                                                    key = k.strip().upper()
                                                    cluster_keys_map[key].add(object_name)
                                        else:
                                            # Fallback to comma-separated string
                                            for k in clustering_cols.split(','):
                                                if k.strip():
                                                    key = k.strip().upper()
                                                    cluster_keys_map[key].add(object_name)
                                    except json.JSONDecodeError:
                                        # If JSON parsing fails, fall back to comma-separated
                                        for k in clustering_cols.split(','):
                                            if k.strip():
                                                key = k.strip().upper()
                                                cluster_keys_map[key].add(object_name)
                    
                    except Exception as e:
                        print(f"⚠️  Error reading {excel_file}: {e}")
                
                if cluster_keys_map:
                    # Prepare output content
                    output_lines = []
                    output_lines.append(f"🔑 Found {len(cluster_keys_map)} unique cluster/partition keys in lineage:")
                    
                    # Analyze query structure to identify JOIN, WHERE, and SELECT sections
                    def analyze_key_usage(key, query_text):
                        """Detect if key is used in JOIN, WHERE, or SELECT clauses"""
                        import re
                        
                        # Convert to uppercase for case-insensitive matching
                        query_upper = query_text.upper()
                        key_pattern = r'\b' + re.escape(key) + r'\b'
                        
                        # Find all matches of the key
                        matches = list(re.finditer(key_pattern, query_upper))
                        
                        in_join = False
                        in_where = False
                        in_select = False
                        
                        for match in matches:
                            pos = match.start()
                            
                            # Get context before the key (up to 500 chars)
                            context_start = max(0, pos - 500)
                            context = query_upper[context_start:pos]
                            
                            # Check if in JOIN clause
                            # Look for JOIN keywords before the key
                            if re.search(r'\b(INNER\s+JOIN|LEFT\s+JOIN|RIGHT\s+JOIN|FULL\s+JOIN|JOIN|LEFT\s+OUTER\s+JOIN|RIGHT\s+OUTER\s+JOIN)\b(?!.*\bWHERE\b)(?!.*\bSELECT\b)', context):
                                # Check if it's in ON clause (after JOIN and before WHERE/GROUP BY/ORDER BY)
                                extended_context = query_upper[context_start:min(len(query_upper), pos + 200)]
                                if re.search(r'\bON\b', extended_context):
                                    in_join = True
                            
                            # Check if in WHERE clause
                            # Look for WHERE keyword before the key and ensure no SELECT after WHERE
                            if re.search(r'\bWHERE\b(?!.*\bSELECT\b)', context):
                                # Make sure we're not in a subquery's SELECT
                                # Check if there's a GROUP BY, ORDER BY, or HAVING after WHERE but before our key
                                remaining_context = query_upper[pos:min(len(query_upper), pos + 100)]
                                # If we see GROUP BY/ORDER BY/HAVING right after, we might be past WHERE clause
                                if not re.search(r'^\s*(GROUP\s+BY|ORDER\s+BY|HAVING|LIMIT|\))', remaining_context):
                                    in_where = True
                            
                            # Check if in SELECT clause
                            # Look for SELECT keyword before the key
                            if re.search(r'\bSELECT\b(?!.*\bFROM\b)', context):
                                # We're in SELECT if SELECT is before key and no FROM between them
                                in_select = True
                            elif re.search(r'\bSELECT\b', context):
                                # Check if we're between SELECT and FROM (select list)
                                from_match = re.search(r'\bFROM\b', context)
                                select_match = re.search(r'\bSELECT\b', context)
                                if select_match and from_match:
                                    # If FROM comes after SELECT, we might be in SELECT list
                                    if select_match.start() > from_match.start():
                                        # SELECT is more recent than FROM, likely in a subquery SELECT
                                        remaining = query_upper[pos:min(len(query_upper), pos + 50)]
                                        if not re.search(r'^\s*FROM\b', remaining):
                                            in_select = True
                        
                        return in_join, in_where, in_select
                    
                    def format_table_list(tables, max_display=5):
                        """Format table list: show first N tables, then '... and X more' if too many"""
                        sorted_tables = sorted(tables)
                        if len(sorted_tables) <= max_display:
                            return ', '.join(sorted_tables)
                        else:
                            displayed = ', '.join(sorted_tables[:max_display])
                            remaining = len(sorted_tables) - max_display
                            return f"{displayed}... and {remaining} more"
                    
                    # Check which keys are present in the query using word boundary matching
                    keys_in_query = []
                    keys_not_in_query = []
                    
                    import re
                    for key in sorted(cluster_keys_map.keys()):
                        # Use word boundary regex to match whole words only
                        pattern = r'\b' + re.escape(key) + r'\b'
                        if re.search(pattern, query_content):
                            keys_in_query.append(key)
                        else:
                            keys_not_in_query.append(key)
                    
                    # Build output for keys in query
                    if keys_in_query:
                        output_lines.append(f"\n✅ Cluster/Partition Keys PRESENT in Query ({len(keys_in_query)}):")
                        for key in keys_in_query:
                            sources = format_table_list(cluster_keys_map[key], max_display=5)
                            in_join, in_where, in_select = analyze_key_usage(key, query_content)
                            usage = f"(Join: {'Y' if in_join else 'N'}, Where: {'Y' if in_where else 'N'}, Select: {'Y' if in_select else 'N'})"
                            output_lines.append(f"   • {key}")
                            output_lines.append(f"     Tables: {sources}")
                            output_lines.append(f"     Usage: {usage}")
                    
                    # Build output for keys not in query
                    if keys_not_in_query:
                        output_lines.append(f"\n⚠️  Cluster/Partition Keys NOT FOUND in Query ({len(keys_not_in_query)}):")
                        for key in keys_not_in_query:
                            sources = format_table_list(cluster_keys_map[key], max_display=3)
                            output_lines.append(f"   • {key}")
                            output_lines.append(f"     Tables: {sources}")
                    
                    # Summary
                    output_lines.append(f"\n📈 Summary:")
                    output_lines.append(f"   Total cluster keys identified: {len(cluster_keys_map)}")
                    output_lines.append(f"   Keys used in query: {len(keys_in_query)} ({len(keys_in_query)*100//len(cluster_keys_map)}%)")
                    output_lines.append(f"   Keys missing from query: {len(keys_not_in_query)} ({len(keys_not_in_query)*100//len(cluster_keys_map)}%)")
                    
                    if len(keys_not_in_query) > 0:
                        output_lines.append(f"\n💡 Recommendation: Consider adding the missing cluster/partition keys to your WHERE clause for better performance")
                    
                    # Write to log file
                    with open(cluster_log_filename, 'w', encoding='utf-8') as log_file:
                        log_file.write('\n'.join(output_lines))
                    
                    # Print to console
                    print('\n'.join(output_lines))
                    print(f"\n📝 Cluster key analysis saved to: {cluster_log_filename}")
                else:
                    output_msg = "ℹ️  No cluster keys or partition keys found in the lineage analysis"
                    print(output_msg)
                    
                    # Write to log file even if no keys found
                    with open(cluster_log_filename, 'w', encoding='utf-8') as log_file:
                        log_file.write(output_msg)
                    print(f"\n📝 Cluster key analysis saved to: {cluster_log_filename}")
                
        except Exception as e:
            print(f"❌ Error analyzing cluster keys: {str(e)}")
            import traceback
            traceback.print_exc()
        
        print("="*80 + "\n")
    
    # Normal mode (direct args or --sql mode)
    else:
        print("\n" + "="*80)
        print("🚀 STARTING LINEAGE ANALYSIS")
        print("="*80)
        print(f"📊 Database: {db}")
        print(f"📊 Schema: {schema}")
        print(f"📊 Objects to process: {len(object_list)}")
        print("\n📋 Object List:")
        for idx, obj in enumerate(object_list, 1):
            print(f"   {idx}. {obj}")
        print("="*80 + "\n")
        
        # Process objects sequentially
        batch_results = process_objects_sequentially(
            database=db,
            schema=schema,
            object_names=object_list
        )


###########################  Unit testing ##################################

# # Test the recursive function
# print("\n" + "="*80)
# print("🧪 TESTING RECURSIVE LINEAGE DISCOVERY")
# print("="*80 + "\n")

# test_results = discover_lineage_recursive(
#     cursor= cursor,
#     database="ZJX_PRD_UHCCD_DB",
#     schema="UDWBASESECUREVIEW1",
#     object_name="ADJD_MCE",
#     level=0
# )

# print("\n" + "="*80)
# print("📊 RESULTS:")
# print("="*80)
# print(f"Total dependencies found: {len(test_results)}")
# print("\nFirst 5 results:")
# import json
# print(json.dumps(test_results[:5], indent=2))



# # Test get_clustering_keys function
# print("\n" + "="*80)
# print("🧪 TESTING get_clustering_keys")
# print("="*80 + "\n")

# # Test with one of your tables
# test_clustering = get_clustering_keys(
#     cursor=cursor,
#     database="ZJX_PRD_UHCCD_DB",
#     table_name="ADJD_MCE"
# )

# print(f"Clustering keys for ADJD_MCE: '{test_clustering}'")

# # Test another table
# test_clustering2 = get_clustering_keys(
#     cursor= cursor,
#     database="ZJX_PRD_UHCCD_DB",
#     table_name="USER_SECUR_KEY_SK"
# )

# print(f"Clustering keys for USER_SECUR_KEY_SK: '{test_clustering2}'")

############################## Test completed ##############################

# test
# test_unauthorized_access
